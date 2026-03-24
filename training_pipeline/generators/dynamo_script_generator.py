"""Training data generator: Dynamo Python node scripts for Revit automation.

Produces ~270 Alpaca-format training pairs covering ProtoGeometry, RevitAPI access
from Dynamo Python nodes, parameter manipulation, family instance operations, list
operations, coordinate transforms, curve/surface operations, and data export.

All output code is Python (Dynamo Python Script node), NOT C#.
Inputs arrive via IN[0], IN[1], ... and results are assigned to OUT.
"""

from __future__ import annotations

from typing import Any, Dict, List

SAMPLE = Dict[str, Any]
MM_TO_FT = 1.0 / 304.8

# Standard Dynamo Python boilerplate snippets reused across samples
_PROTO_HEADER = """\
import clr
clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *"""

_REVIT_HEADER = """\
import clr
clr.AddReference('RevitAPI')
from Autodesk.Revit.DB import *
clr.AddReference('RevitServices')
from RevitServices.Persistence import DocumentManager
from RevitServices.Transactions import TransactionManager
doc = DocumentManager.Instance.CurrentDBDocument"""

_REVIT_NODES_HEADER = """\
import clr
clr.AddReference('RevitNodes')
import Revit
clr.ImportExtensions(Revit.Elements)
clr.ImportExtensions(Revit.GeometryConversion)"""

_TX_START = """\
TransactionManager.Instance.EnsureInTransaction(doc)"""
_TX_END = """\
TransactionManager.Instance.TransactionTaskDone()"""


def _s(instruction: str, output: str) -> SAMPLE:
    return {"instruction": instruction, "input": "", "output": output}


class DynamoScriptGenerator:
    """Generates training samples for Dynamo Python node scripting."""

    def generate(self) -> List[SAMPLE]:
        samples: List[SAMPLE] = []
        samples += self._basic_geometry_creation()
        samples += self._basic_geometry_creation_extra()
        samples += self._revit_element_access()
        samples += self._revit_element_access_extra()
        samples += self._parameter_manipulation()
        samples += self._parameter_manipulation_extra()
        samples += self._family_instance_operations()
        samples += self._family_instance_operations_extra()
        samples += self._list_operations()
        samples += self._list_operations_extra()
        samples += self._coordinate_transforms()
        samples += self._coordinate_transforms_extra()
        samples += self._curve_operations()
        samples += self._curve_operations_extra()
        samples += self._surface_operations()
        samples += self._surface_operations_extra()
        samples += self._export_data()
        samples += self._export_data_extra()
        return samples

    # ------------------------------------------------------------------
    # 1. Basic geometry creation (~40 samples)
    # ------------------------------------------------------------------

    def _basic_geometry_creation(self) -> List[SAMPLE]:
        samples = []

        # Point creation
        point_cases = [
            (0, 0, 0, "at the origin"),
            (1000, 500, 300, "at X=1000, Y=500, Z=300 mm"),
            (0, 0, 2400, "2400 mm above origin"),
        ]
        for x, y, z, desc in point_cases:
            x_ft = x * MM_TO_FT
            y_ft = y * MM_TO_FT
            z_ft = z * MM_TO_FT
            samples.append(_s(
                f"Create a Point {desc} using Dynamo ProtoGeometry",
                f"""\
{_PROTO_HEADER}

# Coordinates in feet (Revit internal units; 1 ft = 304.8 mm)
x = {x_ft:.6f}  # {x} mm
y = {y_ft:.6f}  # {y} mm
z = {z_ft:.6f}  # {z} mm

pt = Point.ByCoordinates(x, y, z)

OUT = pt""",
            ))

        # Line creation
        line_cases = [
            (0, 0, 0, 1000, 0, 0, "horizontal line 1000 mm along X"),
            (0, 0, 0, 0, 0, 3000, "vertical line 3000 mm along Z"),
            (0, 0, 0, 500, 500, 500, "diagonal line to (500,500,500) mm"),
        ]
        for x0, y0, z0, x1, y1, z1, desc in line_cases:
            samples.append(_s(
                f"Create a Line ({desc}) in Dynamo ProtoGeometry",
                f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

start = Point.ByCoordinates({x0 * MM_TO_FT:.6f}, {y0 * MM_TO_FT:.6f}, {z0 * MM_TO_FT:.6f})
end   = Point.ByCoordinates({x1 * MM_TO_FT:.6f}, {y1 * MM_TO_FT:.6f}, {z1 * MM_TO_FT:.6f})

line = Line.ByStartPointEndPoint(start, end)

OUT = line""",
            ))

        # Circle
        circle_cases = [
            (150, "radius 150 mm in the XY plane"),
            (300, "radius 300 mm in the XY plane"),
            (50, "small circle radius 50 mm"),
        ]
        for r, desc in circle_cases:
            samples.append(_s(
                f"Create a Circle ({desc}) in Dynamo ProtoGeometry",
                f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

center_pt = Point.ByCoordinates(0, 0, 0)
normal    = Vector.ByCoordinates(0, 0, 1)
radius    = {r * MM_TO_FT:.6f}  # {r} mm

circle = Circle.ByCenterPointRadiusNormal(center_pt, radius, normal)

OUT = circle""",
            ))

        # Rectangle
        rect_cases = [
            (600, 400, "600x400 mm rectangle"),
            (1200, 900, "1200x900 mm rectangle"),
            (300, 300, "300x300 mm square"),
        ]
        for w, h, desc in rect_cases:
            samples.append(_s(
                f"Create a Rectangle ({desc}) in Dynamo ProtoGeometry",
                f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

origin    = Point.ByCoordinates(0, 0, 0)
x_axis    = Vector.XAxis()
y_axis    = Vector.YAxis()
cs        = CoordinateSystem.ByOriginVectors(origin, x_axis, y_axis)
width     = {w * MM_TO_FT:.6f}  # {w} mm
height    = {h * MM_TO_FT:.6f}  # {h} mm

rect = Rectangle.ByWidthLength(cs, width, height)

OUT = rect""",
            ))

        # PolyCurve from points
        samples.append(_s(
            "Create a PolyCurve from a list of points passed in via IN[0] in Dynamo",
            f"""\
{_PROTO_HEADER}

# IN[0]: list of Point objects
points = IN[0]

poly = PolyCurve.ByPoints(points, connectLastToFirst=False)

OUT = poly""",
        ))

        samples.append(_s(
            "Create a closed PolyCurve (polygon) from a list of XY coordinates in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

# L-shaped polygon in mm, built as list of points
coords_mm = [
    (0,   0),
    (600, 0),
    (600, 300),
    (300, 300),
    (300, 600),
    (0,   600),
]
pts = [Point.ByCoordinates(x * MM_TO_FT, y * MM_TO_FT, 0) for x, y in coords_mm]
poly = PolyCurve.ByPoints(pts, connectLastToFirst=True)

OUT = poly""",
        ))

        # Sphere
        samples.append(_s(
            "Create a Sphere with radius from IN[0] (in mm) at a point from IN[1] in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

radius_mm = IN[0]
center_pt = IN[1]  # expects a Point object; pass None to use origin

radius = radius_mm * MM_TO_FT

if center_pt is None:
    center_pt = Point.ByCoordinates(0, 0, 0)

sphere = Sphere.ByCenterPointRadius(center_pt, radius)

OUT = sphere""",
        ))

        # Box (Cuboid)
        samples.append(_s(
            "Create a Cuboid (box solid) with Width=IN[0], Height=IN[1], Length=IN[2] in mm using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

width_mm  = IN[0]
height_mm = IN[1]
length_mm = IN[2]

cs = CoordinateSystem.Identity()

cuboid = Cuboid.ByLengths(
    cs,
    width_mm  * MM_TO_FT,
    length_mm * MM_TO_FT,
    height_mm * MM_TO_FT,
)

OUT = cuboid""",
        ))

        # Cone
        samples.append(_s(
            "Create a Cone with base radius IN[0] mm, top radius IN[1] mm, and height IN[2] mm in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

base_radius_mm = IN[0]
top_radius_mm  = IN[1]
height_mm      = IN[2]

start_pt = Point.ByCoordinates(0, 0, 0)
end_pt   = Point.ByCoordinates(0, 0, height_mm * MM_TO_FT)
axis     = Line.ByStartPointEndPoint(start_pt, end_pt)

cone = Cone.ByAxisPointsRadii(
    axis,
    start_pt,
    end_pt,
    base_radius_mm * MM_TO_FT,
    top_radius_mm  * MM_TO_FT,
)

OUT = cone""",
        ))

        # Cylinder
        samples.append(_s(
            "Create a Cylinder with radius IN[0] mm and height IN[1] mm along Z in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

radius_mm = IN[0]
height_mm = IN[1]

cs = CoordinateSystem.Identity()
cylinder = Cylinder.ByRadiusHeight(cs, radius_mm * MM_TO_FT, height_mm * MM_TO_FT)

OUT = cylinder""",
        ))

        # Arc
        samples.append(_s(
            "Create an Arc through three points passed in via IN[0], IN[1], IN[2] in Dynamo",
            f"""\
{_PROTO_HEADER}

start_pt  = IN[0]
mid_pt    = IN[1]
end_pt    = IN[2]

arc = Arc.ByThreePoints(start_pt, mid_pt, end_pt)

OUT = arc""",
        ))

        # Ellipse
        samples.append(_s(
            "Create an Ellipse with major radius IN[0] mm and minor radius IN[1] mm in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

major_mm = IN[0]
minor_mm = IN[1]

cs = CoordinateSystem.Identity()
ellipse = Ellipse.ByOriginRadii(
    Point.ByCoordinates(0, 0, 0),
    major_mm * MM_TO_FT,
    minor_mm * MM_TO_FT,
)

OUT = ellipse""",
        ))

        # Bounding box
        samples.append(_s(
            "Compute the bounding box of a geometry object passed in via IN[0] in Dynamo",
            f"""\
{_PROTO_HEADER}

geom = IN[0]

bbox = geom.BoundingBox

min_pt = bbox.MinPoint
max_pt = bbox.MaxPoint

OUT = [bbox, min_pt, max_pt]""",
        ))

        # Solid union
        samples.append(_s(
            "Union two solids passed in via IN[0] and IN[1] using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

solid_a = IN[0]
solid_b = IN[1]

result = solid_a.Union(solid_b)

OUT = result""",
        ))

        # Solid difference
        samples.append(_s(
            "Subtract solid IN[1] from solid IN[0] (boolean difference) in Dynamo",
            f"""\
{_PROTO_HEADER}

base   = IN[0]
cutter = IN[1]

result = base.Difference(cutter)

OUT = result""",
        ))

        # Solid intersection
        samples.append(_s(
            "Intersect two solids passed in via IN[0] and IN[1] in Dynamo",
            f"""\
{_PROTO_HEADER}

solid_a = IN[0]
solid_b = IN[1]

result = solid_a.Intersect(solid_b)

OUT = result""",
        ))

        # Point grid
        samples.append(_s(
            "Generate a grid of points with spacing IN[0] mm, rows IN[1], columns IN[2] in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT  = 1.0 / 304.8
spacing_mm = IN[0]
rows       = IN[1]
cols       = IN[2]

spacing = spacing_mm * MM_TO_FT
pts = []
for r in range(rows):
    row = []
    for c in range(cols):
        row.append(Point.ByCoordinates(c * spacing, r * spacing, 0))
    pts.append(row)

OUT = pts""",
        ))

        # Vector operations
        samples.append(_s(
            "Compute the cross product and dot product of two vectors from IN[0] and IN[1] in Dynamo",
            f"""\
{_PROTO_HEADER}

v1 = IN[0]  # expects Vector
v2 = IN[1]  # expects Vector

cross = v1.Cross(v2)
dot   = v1.Dot(v2)

OUT = [cross, dot]""",
        ))

        # Plane
        samples.append(_s(
            "Create a Plane at origin with normal vector IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

normal = IN[0]  # expects Vector
origin = Point.ByCoordinates(0, 0, 0)

plane = Plane.ByOriginNormal(origin, normal)

OUT = plane""",
        ))

        # Solid from faces (surface thicken)
        samples.append(_s(
            "Thicken a surface passed in via IN[0] by IN[1] mm to produce a solid in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT   = 1.0 / 304.8
surface    = IN[0]
thick_mm   = IN[1]

thickness = thick_mm * MM_TO_FT
solid = surface.Thicken(thickness, both_sides=False)

OUT = solid""",
        ))

        # Point on curve
        samples.append(_s(
            "Get the midpoint and a point at parameter IN[1] (0-1) on a curve IN[0] in Dynamo",
            f"""\
{_PROTO_HEADER}

curve = IN[0]
t     = IN[1]  # parameter in [0, 1]

midpoint     = curve.PointAtParameter(0.5)
point_at_t   = curve.PointAtParameter(t)
tangent_at_t = curve.TangentAtParameter(t)

OUT = [midpoint, point_at_t, tangent_at_t]""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 2. Revit element access (~40 samples)
    # ------------------------------------------------------------------

    def _revit_element_access(self) -> List[SAMPLE]:
        samples = []

        # All walls
        samples.append(_s(
            "Collect all Wall elements in the current Revit document using a Dynamo Python node",
            f"""\
{_REVIT_HEADER}

collector = FilteredElementCollector(doc)
walls = collector.OfClass(Wall).ToElements()

OUT = list(walls)""",
        ))

        # Walls by level
        samples.append(_s(
            "Get all walls on a specific level passed in via IN[0] (Revit Level element) using Dynamo Python",
            f"""\
{_REVIT_HEADER}

level   = UnwrapElement(IN[0])
level_id = level.Id

walls = (FilteredElementCollector(doc)
         .OfClass(Wall)
         .WherePasses(ElementLevelFilter(level_id))
         .ToElements())

OUT = list(walls)""",
        ))

        # Doors
        samples.append(_s(
            "Collect all Door family instances in the active Revit document using Dynamo Python",
            f"""\
{_REVIT_HEADER}

doors = (FilteredElementCollector(doc)
         .OfCategory(BuiltInCategory.OST_Doors)
         .OfClass(FamilyInstance)
         .ToElements())

OUT = list(doors)""",
        ))

        # Rooms
        samples.append(_s(
            "Get all Room elements in the Revit model and return their names and areas using Dynamo Python",
            f"""\
{_REVIT_HEADER}

rooms = (FilteredElementCollector(doc)
         .OfCategory(BuiltInCategory.OST_Rooms)
         .ToElements())

names  = []
areas  = []
for room in rooms:
    names.append(room.get_Parameter(BuiltInParameter.ROOM_NAME).AsString())
    # Area is stored in square feet; convert to m2 for display
    area_sqft = room.get_Parameter(BuiltInParameter.ROOM_AREA).AsDouble()
    areas.append(area_sqft * 0.0929)  # sqft -> m2

OUT = [list(rooms), names, areas]""",
        ))

        # Levels
        samples.append(_s(
            "List all Levels in the project with their elevations in mm using Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8

levels = (FilteredElementCollector(doc)
          .OfClass(Level)
          .ToElements())

names      = [lvl.Name for lvl in levels]
elevations = [lvl.Elevation * FT_TO_MM for lvl in levels]

OUT = [list(levels), names, elevations]""",
        ))

        # Views
        samples.append(_s(
            "Collect all FloorPlan views in the Revit document using a Dynamo Python node",
            f"""\
{_REVIT_HEADER}

views = (FilteredElementCollector(doc)
         .OfClass(View)
         .ToElements())

floor_plans = [v for v in views if v.ViewType == ViewType.FloorPlan]

OUT = floor_plans""",
        ))

        # Structural columns
        samples.append(_s(
            "Get all structural column instances in the model using Dynamo Python",
            f"""\
{_REVIT_HEADER}

columns = (FilteredElementCollector(doc)
           .OfCategory(BuiltInCategory.OST_StructuralColumns)
           .OfClass(FamilyInstance)
           .ToElements())

OUT = list(columns)""",
        ))

        # Floors
        samples.append(_s(
            "Collect all Floor elements and return their types using Dynamo Python",
            f"""\
{_REVIT_HEADER}

floors = (FilteredElementCollector(doc)
          .OfClass(Floor)
          .ToElements())

floor_types = [doc.GetElement(f.GetTypeId()).Name for f in floors]

OUT = [list(floors), floor_types]""",
        ))

        # Windows
        samples.append(_s(
            "Collect all window family instances and return their host wall IDs using Dynamo Python",
            f"""\
{_REVIT_HEADER}

windows = (FilteredElementCollector(doc)
           .OfCategory(BuiltInCategory.OST_Windows)
           .OfClass(FamilyInstance)
           .ToElements())

host_ids = [w.Host.Id.IntegerValue if w.Host is not None else None for w in windows]

OUT = [list(windows), host_ids]""",
        ))

        # Element by ID
        samples.append(_s(
            "Get a Revit element by its integer ID passed in via IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element_id_int = IN[0]
elem_id = ElementId(element_id_int)
element = doc.GetElement(elem_id)

OUT = element""",
        ))

        # Sheets
        samples.append(_s(
            "Get all Sheet views in the project and return their numbers and names using Dynamo Python",
            f"""\
{_REVIT_HEADER}

sheets = (FilteredElementCollector(doc)
          .OfClass(ViewSheet)
          .ToElements())

numbers = [s.SheetNumber for s in sheets]
names   = [s.Name for s in sheets]

OUT = [list(sheets), numbers, names]""",
        ))

        # Family symbols (types)
        samples.append(_s(
            "Get all FamilySymbol types for a specific family by name IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

family_name = IN[0]

families = (FilteredElementCollector(doc)
            .OfClass(Family)
            .ToElements())

target = next((f for f in families if f.Name == family_name), None)

symbols = []
if target is not None:
    for sym_id in target.GetFamilySymbolIds():
        symbols.append(doc.GetElement(sym_id))

OUT = symbols""",
        ))

        # Linked models
        samples.append(_s(
            "Get all Revit Link instances loaded in the document using Dynamo Python",
            f"""\
{_REVIT_HEADER}

links = (FilteredElementCollector(doc)
         .OfClass(RevitLinkInstance)
         .ToElements())

link_names = [lnk.Name for lnk in links]

OUT = [list(links), link_names]""",
        ))

        # MEP pipes
        samples.append(_s(
            "Collect all Pipe elements in the model and return their diameters in mm using Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8

pipes = (FilteredElementCollector(doc)
         .OfCategory(BuiltInCategory.OST_PipeCurves)
         .ToElements())

diameters_mm = []
for pipe in pipes:
    param = pipe.get_Parameter(BuiltInParameter.RBS_PIPE_OUTER_DIAMETER)
    if param is not None:
        diameters_mm.append(param.AsDouble() * FT_TO_MM)
    else:
        diameters_mm.append(None)

OUT = [list(pipes), diameters_mm]""",
        ))

        # Grids
        samples.append(_s(
            "Collect all Grid elements in the project using Dynamo Python",
            f"""\
{_REVIT_HEADER}

grids = (FilteredElementCollector(doc)
         .OfClass(Grid)
         .ToElements())

grid_names = [g.Name for g in grids]

OUT = [list(grids), grid_names]""",
        ))

        # Elements in view
        samples.append(_s(
            "Get all elements visible in the active view using Dynamo Python",
            f"""\
{_REVIT_HEADER}

active_view = doc.ActiveView

elements = (FilteredElementCollector(doc, active_view.Id)
            .ToElements())

OUT = list(elements)""",
        ))

        # Category filter
        samples.append(_s(
            "Collect elements of a specific BuiltInCategory passed as a string in IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

category_name = IN[0]  # e.g. "OST_Furniture"

bic = getattr(BuiltInCategory, category_name)
elems = (FilteredElementCollector(doc)
         .OfCategory(bic)
         .ToElements())

OUT = list(elems)""",
        ))

        # Phases
        samples.append(_s(
            "Get all Phase elements in the project and return their names using Dynamo Python",
            f"""\
{_REVIT_HEADER}

phases = (FilteredElementCollector(doc)
          .OfClass(Phase)
          .ToElements())

phase_names = [p.Name for p in phases]

OUT = [list(phases), phase_names]""",
        ))

        # Materials
        samples.append(_s(
            "Get all Material elements and filter to those whose name contains IN[0] string using Dynamo Python",
            f"""\
{_REVIT_HEADER}

filter_str = IN[0]

materials = (FilteredElementCollector(doc)
             .OfClass(Material)
             .ToElements())

matching = [m for m in materials if filter_str.lower() in m.Name.lower()]

OUT = matching""",
        ))

        # Group instances
        samples.append(_s(
            "Collect all Group instances in the model using Dynamo Python",
            f"""\
{_REVIT_HEADER}

groups = (FilteredElementCollector(doc)
          .OfClass(Group)
          .ToElements())

group_types = [g.GroupType.Name for g in groups]

OUT = [list(groups), group_types]""",
        ))

        # Ceilings
        samples.append(_s(
            "Collect all Ceiling elements and return their type names using Dynamo Python",
            f"""\
{_REVIT_HEADER}

ceilings = (FilteredElementCollector(doc)
            .OfCategory(BuiltInCategory.OST_Ceilings)
            .ToElements())

type_names = [doc.GetElement(c.GetTypeId()).Name for c in ceilings]

OUT = [list(ceilings), type_names]""",
        ))

        # Select elements by workset
        samples.append(_s(
            "Filter all elements belonging to a specific Workset name passed in via IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

workset_name = IN[0]

worksets = (FilteredWorksetCollector(doc)
            .OfKind(WorksetKind.UserWorkset)
            .ToWorksets())

target_ws = next((ws for ws in worksets if ws.Name == workset_name), None)

result = []
if target_ws is not None:
    wsf = ElementWorksetFilter(target_ws.Id)
    result = list(FilteredElementCollector(doc).WherePasses(wsf).ToElements())

OUT = result""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 3. Parameter manipulation (~40 samples)
    # ------------------------------------------------------------------

    def _parameter_manipulation(self) -> List[SAMPLE]:
        samples = []

        # Read a BuiltIn parameter
        samples.append(_s(
            "Read the 'Comments' built-in parameter from a Revit element passed in via IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element = UnwrapElement(IN[0])

param = element.get_Parameter(BuiltInParameter.ALL_MODEL_INSTANCE_COMMENTS)
value = param.AsString() if param is not None else None

OUT = value""",
        ))

        # Read a named parameter
        samples.append(_s(
            "Read the value of a named parameter IN[1] from a Revit element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]

param = element.LookupParameter(param_name)

if param is None:
    value = None
elif param.StorageType == StorageType.Double:
    value = param.AsDouble()
elif param.StorageType == StorageType.Integer:
    value = param.AsInteger()
elif param.StorageType == StorageType.String:
    value = param.AsString()
elif param.StorageType == StorageType.ElementId:
    value = param.AsElementId().IntegerValue
else:
    value = None

OUT = value""",
        ))

        # Set a string parameter
        samples.append(_s(
            "Set the 'Mark' parameter on a Revit element IN[0] to the string value IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
mark_value = IN[1]

{_TX_START}

param = element.get_Parameter(BuiltInParameter.ALL_MODEL_MARK)
if param is not None and not param.IsReadOnly:
    param.Set(str(mark_value))

{_TX_END}

OUT = element""",
        ))

        # Set a length parameter (mm to ft)
        samples.append(_s(
            "Set a length parameter named IN[1] to IN[2] mm on element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

MM_TO_FT = 1.0 / 304.8

element    = UnwrapElement(IN[0])
param_name = IN[1]
value_mm   = IN[2]

{_TX_START}

param = element.LookupParameter(param_name)
if param is not None and not param.IsReadOnly:
    param.Set(value_mm * MM_TO_FT)

{_TX_END}

OUT = element""",
        ))

        # Batch read parameters from list
        samples.append(_s(
            "Read a named parameter from a list of elements IN[0] and return values as a list using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]

values = []
for elem in elements:
    p = elem.LookupParameter(param_name)
    if p is None:
        values.append(None)
    elif p.StorageType == StorageType.Double:
        values.append(p.AsDouble())
    elif p.StorageType == StorageType.Integer:
        values.append(p.AsInteger())
    else:
        values.append(p.AsString())

OUT = values""",
        ))

        # Batch set parameters from list
        samples.append(_s(
            "Set a named parameter IN[1] on each element in list IN[0] to corresponding value in list IN[2] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]
values     = IN[2]

{_TX_START}

results = []
for elem, val in zip(elements, values):
    p = elem.LookupParameter(param_name)
    if p is not None and not p.IsReadOnly:
        if p.StorageType == StorageType.Double:
            p.Set(float(val))
        elif p.StorageType == StorageType.Integer:
            p.Set(int(val))
        else:
            p.Set(str(val))
        results.append(True)
    else:
        results.append(False)

{_TX_END}

OUT = results""",
        ))

        # Unit conversion helper
        samples.append(_s(
            "Convert a Revit internal length value in feet (from AsDouble()) to millimetres in Dynamo Python",
            f"""\
# No Revit API imports required for simple unit math

FT_TO_MM = 304.8

feet_value = IN[0]  # raw AsDouble() result from a length parameter

mm_value = feet_value * FT_TO_MM

OUT = mm_value""",
        ))

        # UnwrapElement usage
        samples.append(_s(
            "Explain and demonstrate the use of UnwrapElement() when working with Dynamo Python nodes",
            f"""\
{_REVIT_HEADER}

# When elements come from Dynamo node wires they are wrapped in Dynamo wrappers.
# UnwrapElement() converts them to raw Revit API objects required by RevitAPI methods.

dynamo_element = IN[0]

# Convert from Dynamo wrapper to Revit API object
revit_element = UnwrapElement(dynamo_element)

# Now safe to call RevitAPI methods:
elem_id   = revit_element.Id
elem_name = Element.Name.GetValue(revit_element)
category  = revit_element.Category.Name

OUT = [elem_id.IntegerValue, elem_name, category]""",
        ))

        # Get all parameters on element
        samples.append(_s(
            "List all parameter names and values on a Revit element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element = UnwrapElement(IN[0])
params  = element.Parameters

names  = []
values = []
for p in params:
    names.append(p.Definition.Name)
    if p.StorageType == StorageType.Double:
        values.append(p.AsDouble())
    elif p.StorageType == StorageType.Integer:
        values.append(p.AsInteger())
    elif p.StorageType == StorageType.String:
        values.append(p.AsString())
    elif p.StorageType == StorageType.ElementId:
        values.append(p.AsElementId().IntegerValue)
    else:
        values.append(None)

OUT = [names, values]""",
        ))

        # Shared parameter check
        samples.append(_s(
            "Check if a parameter named IN[1] on element IN[0] is shared and return its GUID using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]

param = element.LookupParameter(param_name)

is_shared = False
guid_str  = None
if param is not None:
    defn = param.Definition
    if hasattr(defn, 'GUID'):  # ExternalDefinition has GUID; shared params do
        is_shared = True
        guid_str  = str(defn.GUID)

OUT = [is_shared, guid_str]""",
        ))

        # Family type parameter
        samples.append(_s(
            "Read a type parameter from a FamilyInstance IN[0] by name IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance   = UnwrapElement(IN[0])
param_name = IN[1]

symbol = instance.Symbol  # FamilySymbol = type
param  = symbol.LookupParameter(param_name)

value = None
if param is not None:
    if param.StorageType == StorageType.Double:
        value = param.AsDouble()
    elif param.StorageType == StorageType.String:
        value = param.AsString()
    else:
        value = param.AsValueString()

OUT = value""",
        ))

        # Read computed value (AsValueString)
        samples.append(_s(
            "Read the display-formatted value of parameter IN[1] from element IN[0] using AsValueString() in Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]

param = element.LookupParameter(param_name)
display_value = param.AsValueString() if param is not None else None

OUT = display_value""",
        ))

        # Set element name
        samples.append(_s(
            "Rename a Revit element IN[0] to the name string IN[1] (via the Name property) using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element  = UnwrapElement(IN[0])
new_name = IN[1]

{_TX_START}

element.Name = new_name

{_TX_END}

OUT = element""",
        ))

        # Parameter exists check
        samples.append(_s(
            "Check whether a parameter named IN[1] exists on element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]

param  = element.LookupParameter(param_name)
exists = param is not None

OUT = exists""",
        ))

        # Level elevation parameter
        samples.append(_s(
            "Get the elevation of a Level element IN[0] in millimetres using Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8

level = UnwrapElement(IN[0])

elevation_ft = level.Elevation
elevation_mm = elevation_ft * FT_TO_MM

OUT = elevation_mm""",
        ))

        # Wall thickness
        samples.append(_s(
            "Get the Width (thickness) of a Wall element IN[0] in millimetres using Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8

wall = UnwrapElement(IN[0])

# WallType.Width gives the compound thickness
wall_type = doc.GetElement(wall.GetTypeId())
thickness_ft = wall_type.Width
thickness_mm = thickness_ft * FT_TO_MM

OUT = thickness_mm""",
        ))

        # Area parameter
        samples.append(_s(
            "Read the area of a Floor element IN[0] and return it in square metres using Dynamo Python",
            f"""\
{_REVIT_HEADER}

SQFT_TO_SQM = 0.092903

floor = UnwrapElement(IN[0])

area_param = floor.get_Parameter(BuiltInParameter.HOST_AREA_COMPUTED)
area_sqm   = area_param.AsDouble() * SQFT_TO_SQM if area_param is not None else None

OUT = area_sqm""",
        ))

        # Set material parameter by name
        samples.append(_s(
            "Set the material parameter named IN[1] on element IN[0] to the material named IN[2] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element      = UnwrapElement(IN[0])
param_name   = IN[1]
mat_name     = IN[2]

material = next(
    (m for m in FilteredElementCollector(doc).OfClass(Material).ToElements()
     if m.Name == mat_name),
    None
)

if material is not None:
    {_TX_START}
    param = element.LookupParameter(param_name)
    if param is not None and param.StorageType == StorageType.ElementId:
        param.Set(material.Id)
    {_TX_END}

OUT = material""",
        ))

        # Increment a counter parameter
        samples.append(_s(
            "Increment an integer parameter named IN[1] by IN[2] on element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]
increment  = int(IN[2])

param = element.LookupParameter(param_name)

{_TX_START}

if param is not None and param.StorageType == StorageType.Integer:
    current = param.AsInteger()
    param.Set(current + increment)

{_TX_END}

OUT = param.AsInteger() if param is not None else None""",
        ))

        # Read structural usage
        samples.append(_s(
            "Read the Structural Usage parameter from a Wall element IN[0] and return its string value using Dynamo Python",
            f"""\
{_REVIT_HEADER}

wall = UnwrapElement(IN[0])

param = wall.get_Parameter(BuiltInParameter.WALL_STRUCTURAL_USAGE_PARAM)
usage = param.AsValueString() if param is not None else None

OUT = usage""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 4. Family instance operations (~35 samples)
    # ------------------------------------------------------------------

    def _family_instance_operations(self) -> List[SAMPLE]:
        samples = []

        # Place instance at point
        samples.append(_s(
            "Place a FamilyInstance at a point IN[0] using a FamilySymbol IN[1] in Dynamo Python",
            f"""\
{_REVIT_HEADER}

import clr
clr.AddReference('RevitNodes')
import Revit
clr.ImportExtensions(Revit.GeometryConversion)

MM_TO_FT = 1.0 / 304.8

dynamo_point = IN[0]
symbol       = UnwrapElement(IN[1])

# Convert Dynamo Point to Revit XYZ
revit_xyz = dynamo_point.ToRevitType()

{_TX_START}

if not symbol.IsActive:
    symbol.Activate()
    doc.Regenerate()

instance = doc.Create.NewFamilyInstance(
    revit_xyz,
    symbol,
    Structure.StructuralType.NonStructural,
)

{_TX_END}

OUT = instance""",
        ))

        # Place instance on a level
        samples.append(_s(
            "Place a FamilyInstance at X=IN[0], Y=IN[1] mm on Level IN[2] using FamilySymbol IN[3] in Dynamo Python",
            f"""\
{_REVIT_HEADER}

MM_TO_FT = 1.0 / 304.8

x_mm   = IN[0]
y_mm   = IN[1]
level  = UnwrapElement(IN[2])
symbol = UnwrapElement(IN[3])

pt = XYZ(x_mm * MM_TO_FT, y_mm * MM_TO_FT, level.Elevation)

{_TX_START}

if not symbol.IsActive:
    symbol.Activate()
    doc.Regenerate()

instance = doc.Create.NewFamilyInstance(
    pt,
    symbol,
    level,
    Structure.StructuralType.NonStructural,
)

{_TX_END}

OUT = instance""",
        ))

        # Get location point of instance
        samples.append(_s(
            "Get the location point of a FamilyInstance IN[0] and return it as a Dynamo Point using Python",
            f"""\
{_REVIT_HEADER}
{_REVIT_NODES_HEADER}

FT_TO_MM = 304.8

instance = UnwrapElement(IN[0])
loc      = instance.Location

if isinstance(loc, LocationPoint):
    revit_pt = loc.Point
    # Convert to Dynamo Point for use in the graph
    import clr
    clr.AddReference('ProtoGeometry')
    from Autodesk.DesignScript.Geometry import Point as DSPoint
    dynamo_pt = DSPoint.ByCoordinates(
        revit_pt.X * FT_TO_MM,
        revit_pt.Y * FT_TO_MM,
        revit_pt.Z * FT_TO_MM,
    )
    OUT = dynamo_pt
else:
    OUT = None""",
        ))

        # Move instance
        samples.append(_s(
            "Move a FamilyInstance IN[0] by a translation vector (IN[1], IN[2], IN[3]) mm using Dynamo Python",
            f"""\
{_REVIT_HEADER}

MM_TO_FT = 1.0 / 304.8

instance = UnwrapElement(IN[0])
dx_mm    = IN[1]
dy_mm    = IN[2]
dz_mm    = IN[3]

translation = XYZ(dx_mm * MM_TO_FT, dy_mm * MM_TO_FT, dz_mm * MM_TO_FT)

{_TX_START}

ElementTransformUtils.MoveElement(doc, instance.Id, translation)

{_TX_END}

OUT = instance""",
        ))

        # Rotate instance
        samples.append(_s(
            "Rotate a FamilyInstance IN[0] by IN[1] degrees around its location point's Z axis using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import math

instance   = UnwrapElement(IN[0])
degrees    = IN[1]
angle_rad  = degrees * math.pi / 180.0

loc_point = instance.Location.Point
axis = Line.CreateBound(loc_point, loc_point + XYZ.BasisZ)

{_TX_START}

ElementTransformUtils.RotateElement(doc, instance.Id, axis, angle_rad)

{_TX_END}

OUT = instance""",
        ))

        # Change family type
        samples.append(_s(
            "Change the FamilySymbol (type) of a FamilyInstance IN[0] to FamilySymbol IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance   = UnwrapElement(IN[0])
new_symbol = UnwrapElement(IN[1])

{_TX_START}

if not new_symbol.IsActive:
    new_symbol.Activate()
    doc.Regenerate()

instance.Symbol = new_symbol

{_TX_END}

OUT = instance""",
        ))

        # Mirror instance
        samples.append(_s(
            "Mirror a FamilyInstance IN[0] about the YZ plane using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

# Mirror plane: YZ plane (X=0, normal=BasisX)
mirror_plane = Plane.CreateByNormalAndOrigin(XYZ.BasisX, XYZ.Zero)

{_TX_START}

mirrored_ids = ElementTransformUtils.MirrorElements(
    doc,
    [instance.Id],
    mirror_plane,
    mirrored=True,
)

{_TX_END}

mirrored = [doc.GetElement(eid) for eid in mirrored_ids]
OUT = mirrored""",
        ))

        # Copy instance
        samples.append(_s(
            "Copy a FamilyInstance IN[0] to each point in the list IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
{_REVIT_NODES_HEADER}

instance = UnwrapElement(IN[0])
points   = IN[1]  # list of Dynamo Points

origin = instance.Location.Point

{_TX_START}

new_instances = []
for pt in points:
    revit_pt  = pt.ToRevitType()
    translation = revit_pt - origin
    new_ids = ElementTransformUtils.CopyElement(doc, instance.Id, translation)
    new_instances.extend([doc.GetElement(i) for i in new_ids])

{_TX_END}

OUT = new_instances""",
        ))

        # Get host
        samples.append(_s(
            "Return the host element of a FamilyInstance IN[0] (e.g., the wall hosting a door) using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

host = instance.Host  # Wall, Ceiling, Floor etc. or None for unhosted

OUT = host""",
        ))

        # List sub-components
        samples.append(_s(
            "Get the sub-component FamilyInstances of a nested family instance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

sub_ids        = instance.GetSubComponentIds()
sub_components = [doc.GetElement(i) for i in sub_ids]

OUT = sub_components""",
        ))

        # Face-hosted placement
        samples.append(_s(
            "Place a face-based FamilyInstance on a face selected via IN[0] (Reference) using Dynamo Python",
            f"""\
{_REVIT_HEADER}

reference  = IN[0]   # Revit Reference to a face
symbol     = UnwrapElement(IN[1])
work_plane_normal = XYZ.BasisZ

{_TX_START}

if not symbol.IsActive:
    symbol.Activate()
    doc.Regenerate()

instance = doc.Create.NewFamilyInstance(
    reference,
    XYZ.Zero,
    work_plane_normal,
    symbol,
)

{_TX_END}

OUT = instance""",
        ))

        # Get all instance parameters
        samples.append(_s(
            "Return a dictionary of all instance parameter names to values for a FamilyInstance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

param_dict = {{}}
for p in instance.Parameters:
    name = p.Definition.Name
    if p.StorageType == StorageType.Double:
        param_dict[name] = p.AsDouble()
    elif p.StorageType == StorageType.Integer:
        param_dict[name] = p.AsInteger()
    elif p.StorageType == StorageType.String:
        param_dict[name] = p.AsString()
    elif p.StorageType == StorageType.ElementId:
        param_dict[name] = p.AsElementId().IntegerValue

keys   = list(param_dict.keys())
values = list(param_dict.values())

OUT = [keys, values]""",
        ))

        # Delete element
        samples.append(_s(
            "Delete a Revit element IN[0] from the document using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element = UnwrapElement(IN[0])
elem_id = element.Id

{_TX_START}

doc.Delete(elem_id)

{_TX_END}

OUT = elem_id.IntegerValue""",
        ))

        # Place multiple instances from list
        samples.append(_s(
            "Place multiple FamilyInstances from FamilySymbol IN[0] at each point in list IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
{_REVIT_NODES_HEADER}

symbol = UnwrapElement(IN[0])
points = IN[1]  # list of Dynamo Points

{_TX_START}

if not symbol.IsActive:
    symbol.Activate()
    doc.Regenerate()

instances = []
for pt in points:
    revit_pt = pt.ToRevitType()
    inst = doc.Create.NewFamilyInstance(
        revit_pt,
        symbol,
        Structure.StructuralType.NonStructural,
    )
    instances.append(inst)

{_TX_END}

OUT = instances""",
        ))

        # Get facing orientation
        samples.append(_s(
            "Return the facing orientation and hand orientation vectors of a FamilyInstance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

facing = instance.FacingOrientation  # XYZ vector
hand   = instance.HandOrientation    # XYZ vector

OUT = [
    [facing.X, facing.Y, facing.Z],
    [hand.X,   hand.Y,   hand.Z],
]""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 5. List operations (~30 samples)
    # ------------------------------------------------------------------

    def _list_operations(self) -> List[SAMPLE]:
        samples = []

        # Flatten nested list
        samples.append(_s(
            "Flatten a nested list of arbitrary depth passed in via IN[0] using Dynamo Python",
            """\
def flatten(lst):
    result = []
    for item in lst:
        if isinstance(item, list):
            result.extend(flatten(item))
        else:
            result.append(item)
    return result

nested = IN[0]
OUT = flatten(nested)""",
        ))

        # Transpose 2D list
        samples.append(_s(
            "Transpose a 2D list (list of rows) passed in via IN[0] in Dynamo Python",
            """\
matrix = IN[0]

transposed = list(map(list, zip(*matrix)))

OUT = transposed""",
        ))

        # Zip two lists
        samples.append(_s(
            "Zip two lists IN[0] and IN[1] into a list of pairs using Dynamo Python",
            """\
list_a = IN[0]
list_b = IN[1]

pairs = [[a, b] for a, b in zip(list_a, list_b)]

OUT = pairs""",
        ))

        # Filter null / None
        samples.append(_s(
            "Filter None (null) values out of a list IN[0] in Dynamo Python",
            """\
items = IN[0]

filtered = [x for x in items if x is not None]

OUT = filtered""",
        ))

        # Remove duplicates
        samples.append(_s(
            "Remove duplicate values from a list IN[0] while preserving order in Dynamo Python",
            """\
items = IN[0]

seen = set()
unique = []
for item in items:
    if item not in seen:
        seen.add(item)
        unique.append(item)

OUT = unique""",
        ))

        # Sort list of elements by parameter
        samples.append(_s(
            "Sort a list of Revit elements IN[0] by a named parameter IN[1] value using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]

def get_sort_key(elem):
    p = elem.LookupParameter(param_name)
    if p is None:
        return 0
    if p.StorageType == StorageType.Double:
        return p.AsDouble()
    elif p.StorageType == StorageType.Integer:
        return p.AsInteger()
    return p.AsString() or ""

sorted_elems = sorted(elements, key=get_sort_key)

OUT = sorted_elems""",
        ))

        # Chunk list
        samples.append(_s(
            "Split a list IN[0] into sub-lists of length IN[1] (chunking) in Dynamo Python",
            """\
items     = IN[0]
chunk_size = int(IN[1])

chunks = [items[i:i + chunk_size] for i in range(0, len(items), chunk_size)]

OUT = chunks""",
        ))

        # Interleave two lists
        samples.append(_s(
            "Interleave two lists IN[0] and IN[1] element by element in Dynamo Python",
            """\
list_a = IN[0]
list_b = IN[1]

interleaved = []
for pair in zip(list_a, list_b):
    interleaved.extend(pair)

# Append remainder from the longer list
n = min(len(list_a), len(list_b))
interleaved.extend(list_a[n:])
interleaved.extend(list_b[n:])

OUT = interleaved""",
        ))

        # Map (apply a formula to each item)
        samples.append(_s(
            "Multiply every number in a list IN[0] by a factor IN[1] using Dynamo Python",
            """\
numbers = IN[0]
factor  = IN[1]

scaled = [x * factor for x in numbers]

OUT = scaled""",
        ))

        # Group by parameter value
        samples.append(_s(
            "Group a list of Revit elements IN[0] by the value of parameter IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]

groups = {{}}
for elem in elements:
    p = elem.LookupParameter(param_name)
    key = p.AsValueString() if p is not None else "No Parameter"
    groups.setdefault(key, []).append(elem)

keys   = list(groups.keys())
values = list(groups.values())

OUT = [keys, values]""",
        ))

        # Partition list by condition
        samples.append(_s(
            "Split a list IN[0] into two sub-lists: elements where IN[1] parameter value > IN[2] threshold and those where it does not using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]
threshold  = float(IN[2])

above = []
below = []
for elem in elements:
    p = elem.LookupParameter(param_name)
    val = p.AsDouble() if p is not None and p.StorageType == StorageType.Double else 0.0
    (above if val > threshold else below).append(elem)

OUT = [above, below]""",
        ))

        # Create range
        samples.append(_s(
            "Generate a list of evenly spaced numbers from IN[0] to IN[1] with IN[2] steps in Dynamo Python",
            """\
start = float(IN[0])
end   = float(IN[1])
steps = int(IN[2])

if steps < 2:
    OUT = [start]
else:
    step_size = (end - start) / (steps - 1)
    OUT = [start + i * step_size for i in range(steps)]""",
        ))

        # Rotate list
        samples.append(_s(
            "Rotate a list IN[0] by IN[1] positions (like Dynamo's List.Rotate node) in Dynamo Python",
            """\
items  = IN[0]
offset = int(IN[1]) % len(IN[0]) if IN[0] else 0

rotated = items[offset:] + items[:offset]

OUT = rotated""",
        ))

        # Reverse
        samples.append(_s(
            "Reverse the order of a list IN[0] in Dynamo Python",
            """\
items = IN[0]

OUT = list(reversed(items))""",
        ))

        # Get consecutive pairs
        samples.append(_s(
            "Create consecutive pairs from a list IN[0] (e.g., [a,b,c] -> [[a,b],[b,c]]) in Dynamo Python",
            """\
items = IN[0]

pairs = [[items[i], items[i + 1]] for i in range(len(items) - 1)]

OUT = pairs""",
        ))

        # Deduplicate by parameter
        samples.append(_s(
            "Remove duplicate Revit elements IN[0] based on a parameter value IN[1] keeping the first occurrence using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]

seen   = set()
unique = []
for elem in elements:
    p     = elem.LookupParameter(param_name)
    val   = p.AsValueString() if p is not None else str(elem.Id.IntegerValue)
    if val not in seen:
        seen.add(val)
        unique.append(elem)

OUT = unique""",
        ))

        # Flatten and count
        samples.append(_s(
            "Return the total count of leaf items in a nested list IN[0] of any depth using Dynamo Python",
            """\
def count_leaves(lst):
    total = 0
    for item in lst:
        if isinstance(item, list):
            total += count_leaves(item)
        else:
            total += 1
    return total

OUT = count_leaves(IN[0])""",
        ))

        # Combine and deduplicate two element lists
        samples.append(_s(
            "Merge two lists of Revit elements IN[0] and IN[1] and remove duplicates by ElementId using Dynamo Python",
            f"""\
{_REVIT_HEADER}

list_a = [UnwrapElement(e) for e in IN[0]]
list_b = [UnwrapElement(e) for e in IN[1]]

seen = set()
merged = []
for elem in list_a + list_b:
    eid = elem.Id.IntegerValue
    if eid not in seen:
        seen.add(eid)
        merged.append(elem)

OUT = merged""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 6. Coordinate transforms (~30 samples)
    # ------------------------------------------------------------------

    def _coordinate_transforms(self) -> List[SAMPLE]:
        samples = []

        # Identity transform
        samples.append(_s(
            "Create an identity CoordinateSystem at the origin using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

cs = CoordinateSystem.Identity()

OUT = cs""",
        ))

        # CS by origin and axes
        samples.append(_s(
            "Create a CoordinateSystem at point IN[0] with X axis IN[1] and Y axis IN[2] in Dynamo",
            f"""\
{_PROTO_HEADER}

origin = IN[0]  # Point
x_axis = IN[1]  # Vector
y_axis = IN[2]  # Vector

cs = CoordinateSystem.ByOriginVectors(origin, x_axis, y_axis)

OUT = cs""",
        ))

        # Translate geometry
        samples.append(_s(
            "Translate a geometry object IN[0] by vector (IN[1], IN[2], IN[3]) mm using Dynamo Python",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

geom = IN[0]
dx   = IN[1] * MM_TO_FT
dy   = IN[2] * MM_TO_FT
dz   = IN[3] * MM_TO_FT

vec      = Vector.ByCoordinates(dx, dy, dz)
moved    = geom.Translate(vec)

OUT = moved""",
        ))

        # Rotate geometry around Z axis
        samples.append(_s(
            "Rotate a geometry object IN[0] by IN[1] degrees around the Z axis in Dynamo Python",
            f"""\
{_PROTO_HEADER}
import math

geom    = IN[0]
degrees = IN[1]

origin = Point.ByCoordinates(0, 0, 0)
axis   = Vector.ZAxis()

rotated = geom.Rotate(origin, axis, degrees)

OUT = rotated""",
        ))

        # Scale geometry
        samples.append(_s(
            "Scale a geometry object IN[0] uniformly by factor IN[1] about the origin in Dynamo Python",
            f"""\
{_PROTO_HEADER}

geom   = IN[0]
factor = IN[1]

origin = Point.ByCoordinates(0, 0, 0)
cs_from = CoordinateSystem.Identity()
cs_to   = CoordinateSystem.Scale(cs_from, factor)

scaled = geom.Transform(cs_from, cs_to)

OUT = scaled""",
        ))

        # Mirror geometry
        samples.append(_s(
            "Mirror a geometry object IN[0] about the YZ plane (X=0) in Dynamo Python",
            f"""\
{_PROTO_HEADER}

geom = IN[0]

# YZ plane: origin at (0,0,0), normal = X axis
mirror_plane = Plane.ByOriginNormal(
    Point.ByCoordinates(0, 0, 0),
    Vector.XAxis(),
)

mirrored = geom.Mirror(mirror_plane)

OUT = mirrored""",
        ))

        # Transform between two coordinate systems
        samples.append(_s(
            "Transform a geometry object IN[0] from coordinate system IN[1] to coordinate system IN[2] in Dynamo",
            f"""\
{_PROTO_HEADER}

geom   = IN[0]
cs_from = IN[1]  # CoordinateSystem
cs_to   = IN[2]  # CoordinateSystem

transformed = geom.Transform(cs_from, cs_to)

OUT = transformed""",
        ))

        # Project point onto plane
        samples.append(_s(
            "Project a point IN[0] onto a plane IN[1] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

point = IN[0]
plane = IN[1]

projected = point.Project(plane)

OUT = projected""",
        ))

        # Distance between points
        samples.append(_s(
            "Calculate the distance in mm between two points IN[0] and IN[1] (in feet) in Dynamo Python",
            f"""\
{_PROTO_HEADER}

FT_TO_MM = 304.8

pt_a = IN[0]
pt_b = IN[1]

distance_ft = pt_a.DistanceTo(pt_b)
distance_mm = distance_ft * FT_TO_MM

OUT = distance_mm""",
        ))

        # Revit Transform matrix
        samples.append(_s(
            "Create a Revit Transform to move elements by (IN[0], IN[1], IN[2]) mm using Dynamo Python",
            f"""\
{_REVIT_HEADER}

MM_TO_FT = 1.0 / 304.8

dx = IN[0] * MM_TO_FT
dy = IN[1] * MM_TO_FT
dz = IN[2] * MM_TO_FT

t = Transform.CreateTranslation(XYZ(dx, dy, dz))

OUT = t""",
        ))

        # Rotation transform
        samples.append(_s(
            "Create a Revit rotation Transform of IN[0] degrees around the Z axis using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import math

degrees = IN[0]
radians = degrees * math.pi / 180.0

t = Transform.CreateRotation(XYZ.BasisZ, radians)

OUT = t""",
        ))

        # Apply transform to point list
        samples.append(_s(
            "Apply a Revit Transform IN[1] to each XYZ in a list of Revit points IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

transform = IN[1]
points    = IN[0]  # list of XYZ

transformed = [transform.OfPoint(pt) for pt in points]

OUT = transformed""",
        ))

        # Linked model transform
        samples.append(_s(
            "Get the total transform of a RevitLinkInstance IN[0] to convert points from link to host coordinates using Dynamo Python",
            f"""\
{_REVIT_HEADER}

link_instance = UnwrapElement(IN[0])

transform = link_instance.GetTotalTransform()

# Example: transform a point from link coords to host coords
# point_in_host = transform.OfPoint(point_in_link)

OUT = transform""",
        ))

        # Closest point on curve
        samples.append(_s(
            "Find the closest point on a curve IN[0] to a given point IN[1] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

curve = IN[0]
point = IN[1]

closest = curve.ClosestPointTo(point)
param   = curve.ParameterAtPoint(closest)
dist    = point.DistanceTo(closest)

OUT = [closest, param, dist]""",
        ))

        # Remap coordinates between levels
        samples.append(_s(
            "Remap a Z coordinate from Level IN[0] elevation to Level IN[1] elevation in Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8
MM_TO_FT = 1.0 / 304.8

level_from = UnwrapElement(IN[0])
level_to   = UnwrapElement(IN[1])
z_mm       = float(IN[2])

delta_ft = level_to.Elevation - level_from.Elevation
z_remapped_mm = (z_mm * MM_TO_FT + delta_ft) * FT_TO_MM

OUT = z_remapped_mm""",
        ))

        # Array along vector
        samples.append(_s(
            "Array geometry IN[0] IN[1] times along vector IN[2] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

geom   = IN[0]
count  = int(IN[1])
vector = IN[2]  # Vector in feet

copies = []
for i in range(count):
    scaled_vec = vector.Scale(i)
    copies.append(geom.Translate(scaled_vec))

OUT = copies""",
        ))

        # Polar array
        samples.append(_s(
            "Create a polar array of geometry IN[0] with IN[1] copies over IN[2] degrees around Z axis at origin in Dynamo",
            f"""\
{_PROTO_HEADER}

geom    = IN[0]
count   = int(IN[1])
total_deg = float(IN[2])

origin = Point.ByCoordinates(0, 0, 0)
axis   = Vector.ZAxis()
step   = total_deg / count

copies = [geom.Rotate(origin, axis, i * step) for i in range(count)]

OUT = copies""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 7. Curve operations (~30 samples)
    # ------------------------------------------------------------------

    def _curve_operations(self) -> List[SAMPLE]:
        samples = []

        # NurbsCurve by control points
        samples.append(_s(
            "Create a NurbsCurve (degree 3) through a list of control points IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

control_pts = IN[0]  # list of Point objects

nurbs = NurbsCurve.ByControlPoints(control_pts, degree=3)

OUT = nurbs""",
        ))

        # NurbsCurve by points (interpolated)
        samples.append(_s(
            "Create an interpolated NurbsCurve that passes exactly through each point in IN[0] in Dynamo",
            f"""\
{_PROTO_HEADER}

points = IN[0]  # list of Point objects

# ByPoints creates an interpolating spline that passes through all points
nurbs = NurbsCurve.ByPoints(points)

OUT = nurbs""",
        ))

        # Offset curve
        samples.append(_s(
            "Offset a curve IN[0] by IN[1] mm in the XY plane (normal = Z axis) in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

curve    = IN[0]
dist_mm  = IN[1]

dist_ft = dist_mm * MM_TO_FT
normal  = Vector.ByCoordinates(0, 0, 1)

offset = curve.Offset(dist_ft, normal)

OUT = offset""",
        ))

        # Divide curve
        samples.append(_s(
            "Divide a curve IN[0] into IN[1] equal segments and return division points in Dynamo",
            f"""\
{_PROTO_HEADER}

curve = IN[0]
n     = int(IN[1])

params = [i / n for i in range(n + 1)]
pts    = [curve.PointAtParameter(t) for t in params]

OUT = pts""",
        ))

        # Curve length
        samples.append(_s(
            "Get the length of a curve IN[0] in millimetres in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

FT_TO_MM = 304.8

curve = IN[0]

length_ft = curve.Length
length_mm = length_ft * FT_TO_MM

OUT = length_mm""",
        ))

        # Join curves
        samples.append(_s(
            "Join a list of curves IN[0] into a single PolyCurve if they connect end-to-end in Dynamo",
            f"""\
{_PROTO_HEADER}

curves = IN[0]

poly = PolyCurve.ByJoinedCurves(curves)

OUT = poly""",
        ))

        # Explode PolyCurve
        samples.append(_s(
            "Explode a PolyCurve IN[0] into its individual curves in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

poly = IN[0]  # PolyCurve

curves = poly.Curves()

OUT = list(curves)""",
        ))

        # Curve tangent
        samples.append(_s(
            "Get the tangent vector at a parameter IN[1] (0-1) along curve IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

curve = IN[0]
t     = float(IN[1])

tangent = curve.TangentAtParameter(t)
normal  = curve.NormalAtParameter(t)

OUT = [tangent, normal]""",
        ))

        # Trim curve
        samples.append(_s(
            "Trim a curve IN[0] between parameters IN[1] and IN[2] (0-1 range) in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

curve  = IN[0]
t_start = float(IN[1])
t_end   = float(IN[2])

trimmed = curve.ParameterTrim(t_start, t_end)

OUT = trimmed""",
        ))

        # Extend curve
        samples.append(_s(
            "Extend a curve IN[0] at its end by IN[1] mm using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

curve   = IN[0]
ext_mm  = IN[1]
ext_ft  = ext_mm * MM_TO_FT

extended = curve.Extend(0, ext_ft)

OUT = extended""",
        ))

        # Intersection of two curves
        samples.append(_s(
            "Find the intersection point(s) of two curves IN[0] and IN[1] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

curve_a = IN[0]
curve_b = IN[1]

intersections = curve_a.Intersect(curve_b)

OUT = list(intersections)""",
        ))

        # Helix
        samples.append(_s(
            "Create a helix with radius IN[0] mm, pitch IN[1] mm per turn, and IN[2] turns in Dynamo",
            f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

radius_mm = IN[0]
pitch_mm  = IN[1]
turns     = IN[2]

radius = radius_mm * MM_TO_FT
pitch  = pitch_mm  * MM_TO_FT

n_pts = int(turns * 36) + 1  # 36 points per turn
pts = []
for i in range(n_pts):
    t     = i / 36.0  # turns
    angle = 2 * math.pi * t
    z     = pitch * t
    pts.append(Point.ByCoordinates(
        radius * math.cos(angle),
        radius * math.sin(angle),
        z,
    ))

helix = NurbsCurve.ByPoints(pts)

OUT = helix""",
        ))

        # Fillet two lines
        samples.append(_s(
            "Create a fillet arc between two intersecting lines IN[0] and IN[1] with radius IN[2] mm in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

line_a  = IN[0]
line_b  = IN[1]
rad_mm  = IN[2]
rad_ft  = rad_mm * MM_TO_FT

fillet = PolyCurve.ByFillet(PolyCurve.ByJoinedCurves([line_a, line_b]), rad_ft)

OUT = fillet""",
        ))

        # Parallel curves
        samples.append(_s(
            "Generate IN[1] parallel offset curves from curve IN[0] spaced IN[2] mm apart in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

curve   = IN[0]
count   = int(IN[1])
gap_mm  = IN[2]

normal  = Vector.ZAxis()
offsets = []
for i in range(count):
    dist = i * gap_mm * MM_TO_FT
    offsets.append(curve.Offset(dist, normal))

OUT = offsets""",
        ))

        # Closest point on curve to point
        samples.append(_s(
            "Get the closest point on curve IN[0] to point IN[1] and return the distance in mm in Dynamo",
            f"""\
{_PROTO_HEADER}

FT_TO_MM = 304.8

curve = IN[0]
pt    = IN[1]

closest    = curve.ClosestPointTo(pt)
dist_ft    = pt.DistanceTo(closest)
dist_mm    = dist_ft * FT_TO_MM

OUT = [closest, dist_mm]""",
        ))

        # Reverse curve direction
        samples.append(_s(
            "Reverse the direction of a curve IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

curve = IN[0]

reversed_curve = curve.Reverse()

OUT = reversed_curve""",
        ))

        # Degree elevation NurbsCurve
        samples.append(_s(
            "Create a degree-5 NurbsCurve from a list of weighted control points IN[0] in Dynamo",
            f"""\
{_PROTO_HEADER}

control_pts = IN[0]  # list of Point
weights     = IN[1]  # list of floats, same length as control_pts; pass None to use uniform weights

degree = 5

if weights is None:
    nurbs = NurbsCurve.ByControlPoints(control_pts, degree)
else:
    nurbs = NurbsCurve.ByControlPointsWeightsKnots(
        control_pts,
        weights,
        None,   # let Dynamo compute uniform knots
        degree,
    )

OUT = nurbs""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 8. Surface operations (~25 samples)
    # ------------------------------------------------------------------

    def _surface_operations(self) -> List[SAMPLE]:
        samples = []

        # Loft surface
        samples.append(_s(
            "Create a lofted Surface through a list of section curves IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

profiles = IN[0]  # list of curves (e.g., Lines, NurbsCurves)

surface = Surface.ByLoft(profiles)

OUT = surface""",
        ))

        # Patch (cap planar boundary)
        samples.append(_s(
            "Create a planar Surface that patches a closed curve IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

closed_curve = IN[0]

surface = Surface.ByPatch(closed_curve)

OUT = surface""",
        ))

        # Revolve surface
        samples.append(_s(
            "Create a surface of revolution from profile curve IN[0] around the Z axis by IN[1] degrees in Dynamo",
            f"""\
{_PROTO_HEADER}

profile = IN[0]
degrees = float(IN[1])

axis_origin = Point.ByCoordinates(0, 0, 0)
axis_dir    = Vector.ZAxis()

surface = Surface.ByRevolve(profile, axis_origin, axis_dir, 0, degrees)

OUT = surface""",
        ))

        # Surface area
        samples.append(_s(
            "Calculate the area of a surface IN[0] and return it in mm^2 using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

surface = IN[0]

FT2_TO_MM2 = 304.8 * 304.8

area_ft2 = surface.Area
area_mm2 = area_ft2 * FT2_TO_MM2

OUT = area_mm2""",
        ))

        # Surface normal at UV
        samples.append(_s(
            "Get the surface normal at parameter (IN[1], IN[2]) on surface IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

surface = IN[0]
u       = float(IN[1])
v       = float(IN[2])

normal   = surface.NormalAtParameter(u, v)
point    = surface.PointAtParameter(u, v)

OUT = [point, normal]""",
        ))

        # Offset surface
        samples.append(_s(
            "Offset a surface IN[0] by IN[1] mm along its normals in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

surface = IN[0]
dist_mm = IN[1]

offset = surface.Offset(dist_mm * MM_TO_FT)

OUT = offset""",
        ))

        # Thicken to solid
        samples.append(_s(
            "Thicken surface IN[0] by IN[1] mm on both sides to produce a solid in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

surface  = IN[0]
thick_mm = IN[1]

solid = surface.Thicken(thick_mm * MM_TO_FT, both_sides=True)

OUT = solid""",
        ))

        # UV point grid on surface
        samples.append(_s(
            "Sample a grid of IN[1] x IN[2] points on surface IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

surface = IN[0]
u_count = int(IN[1])
v_count = int(IN[2])

pts = []
for i in range(u_count):
    row = []
    u = i / (u_count - 1) if u_count > 1 else 0.5
    for j in range(v_count):
        v = j / (v_count - 1) if v_count > 1 else 0.5
        row.append(surface.PointAtParameter(u, v))
    pts.append(row)

OUT = pts""",
        ))

        # Surface by extrude
        samples.append(_s(
            "Extrude a curve IN[0] by a direction vector IN[1] (in mm) to produce a Surface in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

curve     = IN[0]
direction = IN[1]  # Vector in mm

scaled_dir = direction.Scale(MM_TO_FT)
surface    = Surface.ByExtrude(curve, scaled_dir)

OUT = surface""",
        ))

        # Solid from closed surfaces
        samples.append(_s(
            "Create a Solid from a list of closed surfaces IN[0] by stitching them together in Dynamo",
            f"""\
{_PROTO_HEADER}

surfaces = IN[0]  # list of Surface objects forming a closed shell

solid = Solid.ByJoinedSurfaces(surfaces)

OUT = solid""",
        ))

        # Planar surface by rectangle
        samples.append(_s(
            "Create a rectangular planar Surface with width IN[0] mm and height IN[1] mm in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

width_mm  = IN[0]
height_mm = IN[1]

w = width_mm  * MM_TO_FT
h = height_mm * MM_TO_FT

p0 = Point.ByCoordinates(-w/2, -h/2, 0)
p1 = Point.ByCoordinates( w/2, -h/2, 0)
p2 = Point.ByCoordinates( w/2,  h/2, 0)
p3 = Point.ByCoordinates(-w/2,  h/2, 0)

boundary = PolyCurve.ByPoints([p0, p1, p2, p3], connectLastToFirst=True)
surface  = Surface.ByPatch(boundary)

OUT = surface""",
        ))

        # Intersect surface with plane
        samples.append(_s(
            "Intersect surface IN[0] with a horizontal plane at Z=IN[1] mm to get cross-section curves in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

surface = IN[0]
z_mm    = IN[1]

cut_plane = Plane.ByOriginNormal(
    Point.ByCoordinates(0, 0, z_mm * MM_TO_FT),
    Vector.ZAxis(),
)

curves = surface.Intersect(cut_plane)

OUT = list(curves)""",
        ))

        # Surface perimeter
        samples.append(_s(
            "Get the perimeter curves of a surface IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

surface = IN[0]

perimeter_curves = surface.PerimeterCurves()

OUT = list(perimeter_curves)""",
        ))

        # Sweep surface
        samples.append(_s(
            "Sweep a profile curve IN[0] along a path curve IN[1] to create a surface in Dynamo",
            f"""\
{_PROTO_HEADER}

profile = IN[0]  # cross-section curve
path    = IN[1]  # sweep path curve

surface = Surface.BySweep(profile, path)

OUT = surface""",
        ))

        # Bi-rail sweep
        samples.append(_s(
            "Create a bi-rail sweep surface using profile IN[0] along two rail curves IN[1] and IN[2] in Dynamo",
            f"""\
{_PROTO_HEADER}

profile = IN[0]
rail_a  = IN[1]
rail_b  = IN[2]

surface = Surface.BySweep2Rails(rail_a, rail_b, profile)

OUT = surface""",
        ))

        return samples

    # ------------------------------------------------------------------
    # 9. Export data (~30 samples)
    # ------------------------------------------------------------------

    def _export_data(self) -> List[SAMPLE]:
        samples = []

        # Write to CSV
        samples.append(_s(
            "Write a list of rows IN[0] (list of lists) to a CSV file at path IN[1] using Dynamo Python",
            """\
import csv
import os

rows      = IN[0]  # list of lists
file_path = IN[1]  # full file path string, e.g. r"C:\\output\\data.csv"

os.makedirs(os.path.dirname(file_path), exist_ok=True)

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerows(rows)

OUT = file_path""",
        ))

        # Write with header
        samples.append(_s(
            "Write a CSV with a header row IN[0] and data rows IN[1] to file IN[2] using Dynamo Python",
            """\
import csv

header    = IN[0]  # list of column names
data_rows = IN[1]  # list of lists
file_path = IN[2]

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(data_rows)

OUT = file_path""",
        ))

        # Append to CSV
        samples.append(_s(
            "Append a new row IN[0] to an existing CSV file IN[1] using Dynamo Python",
            """\
import csv

new_row   = IN[0]
file_path = IN[1]

with open(file_path, 'a', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(new_row)

OUT = file_path""",
        ))

        # Read CSV
        samples.append(_s(
            "Read all rows from a CSV file at path IN[0] and return as a list of lists using Dynamo Python",
            """\
import csv

file_path = IN[0]

rows = []
with open(file_path, 'r', newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row in reader:
        rows.append(row)

header    = rows[0] if rows else []
data_rows = rows[1:] if len(rows) > 1 else []

OUT = [header, data_rows]""",
        ))

        # Write to Excel with openpyxl
        samples.append(_s(
            "Write a 2D list IN[0] to an Excel file at path IN[1] using openpyxl in Dynamo Python",
            """\
# Requires openpyxl: pip install openpyxl (install in Dynamo's Python environment)
import clr
import sys

# Ensure openpyxl is on the path if installed separately
# sys.path.append(r"C:\\path\\to\\site-packages")
import openpyxl

data      = IN[0]  # list of lists
file_path = IN[1]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

for row in data:
    ws.append(row)

wb.save(file_path)

OUT = file_path""",
        ))

        # Read Excel with openpyxl
        samples.append(_s(
            "Read all values from the first sheet of an Excel file IN[0] and return as a 2D list using Dynamo Python",
            """\
import openpyxl

file_path = IN[0]

wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

rows = []
for row in ws.iter_rows(values_only=True):
    rows.append(list(row))

OUT = rows""",
        ))

        # Write JSON
        samples.append(_s(
            "Serialize a Python dictionary or list IN[0] to a JSON file at path IN[1] using Dynamo Python",
            """\
import json

data      = IN[0]
file_path = IN[1]

with open(file_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

OUT = file_path""",
        ))

        # Read JSON
        samples.append(_s(
            "Read and deserialize a JSON file at path IN[0] using Dynamo Python",
            """\
import json

file_path = IN[0]

with open(file_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

OUT = data""",
        ))

        # Export element parameters to CSV
        samples.append(_s(
            "Export parameter values of a list of elements IN[0] for parameter names IN[1] to CSV at path IN[2] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

elements    = [UnwrapElement(e) for e in IN[0]]
param_names = IN[1]  # list of parameter name strings
file_path   = IN[2]

rows = [param_names]  # header
for elem in elements:
    row = []
    for name in param_names:
        p = elem.LookupParameter(name)
        if p is None:
            row.append('')
        elif p.StorageType == StorageType.Double:
            row.append(p.AsDouble())
        elif p.StorageType == StorageType.Integer:
            row.append(p.AsInteger())
        else:
            row.append(p.AsValueString() or '')
    rows.append(row)

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(rows)

OUT = file_path""",
        ))

        # Read schedule data
        samples.append(_s(
            "Read data from a Revit ViewSchedule IN[0] and return it as a 2D list using Dynamo Python",
            f"""\
{_REVIT_HEADER}

schedule = UnwrapElement(IN[0])  # ViewSchedule element

table  = schedule.GetTableData()
body   = table.GetSectionData(SectionType.Body)
header = table.GetSectionData(SectionType.Header)

n_rows = body.NumberOfRows
n_cols = body.NumberOfColumns

rows = []
for r in range(n_rows):
    row = [schedule.GetCellText(SectionType.Body, r, c) for c in range(n_cols)]
    rows.append(row)

header_row = [schedule.GetCellText(SectionType.Header, 0, c) for c in range(n_cols)]

OUT = [header_row, rows]""",
        ))

        # Write text file
        samples.append(_s(
            "Write a multi-line text report from list IN[0] to a .txt file at path IN[1] using Dynamo Python",
            """\
file_path = IN[1]
lines     = [str(item) for item in IN[0]]

with open(file_path, 'w', encoding='utf-8') as f:
    f.write('\\n'.join(lines))

OUT = file_path""",
        ))

        # Read text file lines
        samples.append(_s(
            "Read all lines from a text file at path IN[0] and return as a list using Dynamo Python",
            """\
file_path = IN[0]

with open(file_path, 'r', encoding='utf-8') as f:
    lines = [line.rstrip('\\n') for line in f.readlines()]

OUT = lines""",
        ))

        # Export to Excel with named sheet
        samples.append(_s(
            "Write data IN[0] to a named worksheet IN[1] in an Excel file IN[2] using openpyxl in Dynamo Python",
            """\
import openpyxl
import os

data       = IN[0]
sheet_name = IN[1]
file_path  = IN[2]

if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path)
else:
    wb = openpyxl.Workbook()
    # remove default sheet if adding a named one
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)  # clear existing data
else:
    ws = wb.create_sheet(title=sheet_name)

for row in data:
    ws.append([str(v) for v in row])

wb.save(file_path)

OUT = file_path""",
        ))

        # Export wall schedule
        samples.append(_s(
            "Export all Wall elements with their length, height, and type name to CSV at path IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

FT_TO_MM = 304.8
file_path = IN[0]

walls = (FilteredElementCollector(doc)
         .OfClass(Wall)
         .ToElements())

header = ['ElementId', 'TypeName', 'Length_mm', 'Height_mm', 'Volume_m3']
rows   = [header]

for wall in walls:
    eid       = wall.Id.IntegerValue
    type_name = doc.GetElement(wall.GetTypeId()).Name
    length_p  = wall.get_Parameter(BuiltInParameter.CURVE_ELEM_LENGTH)
    height_p  = wall.get_Parameter(BuiltInParameter.WALL_USER_HEIGHT_PARAM)
    vol_p     = wall.get_Parameter(BuiltInParameter.HOST_VOLUME_COMPUTED)
    length_mm = length_p.AsDouble() * FT_TO_MM if length_p else ''
    height_mm = height_p.AsDouble() * FT_TO_MM if height_p else ''
    vol_m3    = vol_p.AsDouble() * 0.0283168 if vol_p else ''  # cuft to m3
    rows.append([eid, type_name, length_mm, height_mm, vol_m3])

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(rows)

OUT = file_path""",
        ))

        # Export room data
        samples.append(_s(
            "Export all Room elements (name, number, area m2, level) to CSV at path IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

SQFT_TO_SQM = 0.092903
file_path   = IN[0]

rooms = (FilteredElementCollector(doc)
         .OfCategory(BuiltInCategory.OST_Rooms)
         .ToElements())

header = ['ElementId', 'RoomName', 'RoomNumber', 'Area_m2', 'Level']
rows   = [header]

for room in rooms:
    eid    = room.Id.IntegerValue
    name_p = room.get_Parameter(BuiltInParameter.ROOM_NAME)
    num_p  = room.get_Parameter(BuiltInParameter.ROOM_NUMBER)
    area_p = room.get_Parameter(BuiltInParameter.ROOM_AREA)
    level  = doc.GetElement(room.LevelId).Name if room.LevelId != ElementId.InvalidElementId else ''
    rows.append([
        eid,
        name_p.AsString() if name_p else '',
        num_p.AsString()  if num_p  else '',
        area_p.AsDouble() * SQFT_TO_SQM if area_p else '',
        level,
    ])

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(rows)

OUT = file_path""",
        ))

        # Read CSV and set parameters
        samples.append(_s(
            "Read a CSV file IN[0] with columns 'ElementId' and a parameter name IN[1] and set those values using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

file_path  = IN[0]
param_name = IN[1]

with open(file_path, 'r', newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    rows   = list(reader)

{_TX_START}

results = []
for row in rows:
    eid  = ElementId(int(row['ElementId']))
    elem = doc.GetElement(eid)
    if elem is None:
        results.append(False)
        continue
    p = elem.LookupParameter(param_name)
    if p is not None and not p.IsReadOnly:
        p.Set(str(row.get(param_name, '')))
        results.append(True)
    else:
        results.append(False)

{_TX_END}

OUT = results""",
        ))

        # Print to Dynamo watch node
        samples.append(_s(
            "Format a list of elements IN[0] as human-readable strings for display in a Dynamo Watch node using Python",
            f"""\
{_REVIT_HEADER}

elements = [UnwrapElement(e) for e in IN[0]]

lines = []
for elem in elements:
    eid      = elem.Id.IntegerValue
    cat_name = elem.Category.Name if elem.Category else 'No Category'
    try:
        name = elem.Name
    except Exception:
        name = '(no name)'
    lines.append(f"{{eid}} | {{cat_name}} | {{name}}")

OUT = lines""",
        ))

        # Check file exists
        samples.append(_s(
            "Check whether a file path IN[0] exists and return file size in KB using Dynamo Python",
            """\
import os

file_path = IN[0]

exists    = os.path.exists(file_path)
size_kb   = os.path.getsize(file_path) / 1024.0 if exists else 0

OUT = [exists, size_kb]""",
        ))

        # List files in directory
        samples.append(_s(
            "List all files with extension IN[1] in directory IN[0] using Dynamo Python",
            """\
import os

directory = IN[0]
extension = IN[1]  # e.g. ".rvt"

files = [
    os.path.join(directory, f)
    for f in os.listdir(directory)
    if f.lower().endswith(extension.lower())
]

OUT = sorted(files)""",
        ))

        # Log with timestamp
        samples.append(_s(
            "Append a timestamped log entry IN[0] to a log file IN[1] using Dynamo Python",
            """\
import datetime

message   = IN[0]
file_path = IN[1]

timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
log_line  = f"[{timestamp}] {message}\\n"

with open(file_path, 'a', encoding='utf-8') as f:
    f.write(log_line)

OUT = log_line""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- basic geometry creation
    # ------------------------------------------------------------------

    def _basic_geometry_creation_extra(self) -> List[SAMPLE]:
        samples = []

        # Torus
        samples.append(_s(
            "Create a Torus with major radius IN[0] mm and minor (tube) radius IN[1] mm in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

major_mm = IN[0]
minor_mm = IN[1]

major = major_mm * MM_TO_FT
minor = minor_mm * MM_TO_FT

# Approximate torus as a revolution of a small circle around Z
small_circle_center = Point.ByCoordinates(major, 0, 0)
small_circle_normal = Vector.YAxis()
profile = Circle.ByCenterPointRadiusNormal(small_circle_center, minor, small_circle_normal)

axis_origin = Point.ByCoordinates(0, 0, 0)
axis_dir    = Vector.ZAxis()

torus = Surface.ByRevolve(profile, axis_origin, axis_dir, 0, 360)

OUT = torus""",
        ))

        # RegularPolygon approximate via PolyCurve
        for n_sides, desc in [(3, "equilateral triangle"), (6, "regular hexagon"), (8, "regular octagon")]:
            samples.append(_s(
                f"Create a regular {n_sides}-sided polygon ({desc}) inscribed in a circle of radius IN[0] mm in Dynamo",
                f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

radius_mm = IN[0]
radius    = radius_mm * MM_TO_FT
n_sides   = {n_sides}

pts = [
    Point.ByCoordinates(
        radius * math.cos(2 * math.pi * i / n_sides),
        radius * math.sin(2 * math.pi * i / n_sides),
        0,
    )
    for i in range(n_sides)
]

poly = PolyCurve.ByPoints(pts, connectLastToFirst=True)

OUT = poly""",
            ))

        # Pipe (hollow cylinder as two-surface solid)
        samples.append(_s(
            "Create a hollow pipe solid with outer radius IN[0] mm, inner radius IN[1] mm, length IN[2] mm in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

outer_mm = IN[0]
inner_mm = IN[1]
len_mm   = IN[2]

outer_cyl = Cylinder.ByRadiusHeight(
    CoordinateSystem.Identity(),
    outer_mm * MM_TO_FT,
    len_mm   * MM_TO_FT,
)
inner_cyl = Cylinder.ByRadiusHeight(
    CoordinateSystem.Identity(),
    inner_mm * MM_TO_FT,
    len_mm   * MM_TO_FT,
)

pipe = outer_cyl.Difference(inner_cyl)

OUT = pipe""",
        ))

        # Random point cloud
        samples.append(_s(
            "Generate IN[0] random points within a bounding box of size IN[1] x IN[2] x IN[3] mm in Dynamo Python",
            f"""\
{_PROTO_HEADER}
import random

MM_TO_FT = 1.0 / 304.8

count  = int(IN[0])
x_mm   = IN[1]
y_mm   = IN[2]
z_mm   = IN[3]

pts = [
    Point.ByCoordinates(
        random.uniform(0, x_mm) * MM_TO_FT,
        random.uniform(0, y_mm) * MM_TO_FT,
        random.uniform(0, z_mm) * MM_TO_FT,
    )
    for _ in range(count)
]

OUT = pts""",
        ))

        # Convex hull placeholder note
        samples.append(_s(
            "Create a mesh from a list of points IN[0] using Delaunay triangulation in Dynamo Python (using MeshToolkit if available)",
            f"""\
{_PROTO_HEADER}

# MeshToolkit must be installed as a Dynamo package for Mesh.ByGeometry
# Fallback: create a NurbsCurve hull approximation

points = IN[0]  # list of Point objects

# If MeshToolkit available:
# clr.AddReference('MeshToolkit')
# from Autodesk.Dynamo.MeshToolkit import Mesh
# mesh = Mesh.ByVerticesFaceIndices(points, face_indices)

# Without MeshToolkit -- return a bounding convex hull via surface loft approximation
sorted_pts = sorted(points, key=lambda p: (p.X, p.Y))
hull_curve = NurbsCurve.ByPoints(sorted_pts + [sorted_pts[0]])

OUT = hull_curve""",
        ))

        # Point from polar coords
        samples.append(_s(
            "Create a Point from polar coordinates (radius IN[0] mm, angle IN[1] degrees, Z IN[2] mm) in Dynamo Python",
            f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

r_mm    = IN[0]
deg     = IN[1]
z_mm    = IN[2]

r = r_mm * MM_TO_FT
z = z_mm * MM_TO_FT

rad = math.radians(deg)
x   = r * math.cos(rad)
y   = r * math.sin(rad)

pt = Point.ByCoordinates(x, y, z)

OUT = pt""",
        ))

        # Solid from curves via sweep
        samples.append(_s(
            "Sweep a circular profile of radius IN[0] mm along a path PolyCurve IN[1] to make a pipe solid in Dynamo",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

radius_mm = IN[0]
path      = IN[1]  # PolyCurve or Curve

# Profile circle at path start point, in the plane perpendicular to the path tangent
start_pt      = path.StartPoint
tangent_start = path.TangentAtParameter(0)
profile_cs    = CoordinateSystem.ByOriginVectors(
    start_pt,
    tangent_start.Cross(Vector.ZAxis()).Normalized(),
    Vector.ZAxis(),
)

profile = Circle.ByCenterPointRadiusNormal(
    start_pt,
    radius_mm * MM_TO_FT,
    tangent_start,
)

solid = Solid.BySweep(profile, path)

OUT = solid""",
        ))

        # Revolved solid
        samples.append(_s(
            "Create a solid of revolution by revolving profile curve IN[0] around the Y axis by 360 degrees in Dynamo",
            f"""\
{_PROTO_HEADER}

profile = IN[0]  # open curve in the XZ plane

axis_origin = Point.ByCoordinates(0, 0, 0)
axis_dir    = Vector.YAxis()

solid = Solid.ByRevolve(profile, axis_origin, axis_dir, 0, 360)

OUT = solid""",
        ))

        # Signed area of polygon
        samples.append(_s(
            "Calculate the signed area of a 2D polygon defined by a list of XY coordinate pairs IN[0] using Dynamo Python (shoelace formula)",
            """\
# Shoelace formula -- no Revit API needed
# IN[0]: list of [x, y] coordinate pairs in any consistent unit

coords = IN[0]
n = len(coords)

area = 0.0
for i in range(n):
    x0, y0 = coords[i]
    x1, y1 = coords[(i + 1) % n]
    area += x0 * y1 - x1 * y0

signed_area = area / 2.0
abs_area    = abs(signed_area)

OUT = [signed_area, abs_area]""",
        ))

        # Grid shell point array
        samples.append(_s(
            "Generate a sinusoidal surface point grid with IN[0] x IN[1] points, spacing IN[2] mm, amplitude IN[3] mm in Dynamo",
            f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

rows      = int(IN[0])
cols      = int(IN[1])
spacing_mm = IN[2]
amp_mm    = IN[3]

spacing = spacing_mm * MM_TO_FT
amp     = amp_mm     * MM_TO_FT

pts = []
for r in range(rows):
    row = []
    for c in range(cols):
        x = c * spacing
        y = r * spacing
        z = amp * math.sin(2 * math.pi * c / cols) * math.cos(2 * math.pi * r / rows)
        row.append(Point.ByCoordinates(x, y, z))
    pts.append(row)

OUT = pts""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- Revit element access
    # ------------------------------------------------------------------

    def _revit_element_access_extra(self) -> List[SAMPLE]:
        samples = []

        # MEP ducts
        samples.append(_s(
            "Collect all Duct elements in the model and return their system types using Dynamo Python",
            f"""\
{_REVIT_HEADER}

ducts = (FilteredElementCollector(doc)
         .OfCategory(BuiltInCategory.OST_DuctCurves)
         .ToElements())

system_types = []
for duct in ducts:
    p = duct.get_Parameter(BuiltInParameter.RBS_DUCT_SYSTEM_TYPE_PARAM)
    system_types.append(p.AsValueString() if p else None)

OUT = [list(ducts), system_types]""",
        ))

        # Furniture
        samples.append(_s(
            "Collect all Furniture family instances and return their family and type names using Dynamo Python",
            f"""\
{_REVIT_HEADER}

furniture = (FilteredElementCollector(doc)
             .OfCategory(BuiltInCategory.OST_Furniture)
             .OfClass(FamilyInstance)
             .ToElements())

family_names = [f.Symbol.Family.Name for f in furniture]
type_names   = [f.Symbol.Name for f in furniture]

OUT = [list(furniture), family_names, type_names]""",
        ))

        # Annotation tags
        samples.append(_s(
            "Collect all independent tag elements in the active view using Dynamo Python",
            f"""\
{_REVIT_HEADER}

active_view = doc.ActiveView

tags = (FilteredElementCollector(doc, active_view.Id)
        .OfClass(IndependentTag)
        .ToElements())

OUT = list(tags)""",
        ))

        # Stairs
        samples.append(_s(
            "Get all Stair elements in the model using Dynamo Python",
            f"""\
{_REVIT_HEADER}

stairs = (FilteredElementCollector(doc)
          .OfCategory(BuiltInCategory.OST_Stairs)
          .ToElements())

OUT = list(stairs)""",
        ))

        # Curtain panels
        samples.append(_s(
            "Get all Curtain Panel family instances in the model using Dynamo Python",
            f"""\
{_REVIT_HEADER}

panels = (FilteredElementCollector(doc)
          .OfCategory(BuiltInCategory.OST_CurtainWallPanels)
          .OfClass(FamilyInstance)
          .ToElements())

OUT = list(panels)""",
        ))

        # Railings
        samples.append(_s(
            "Collect all Railing elements and return their top rail height in mm using Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8

railings = (FilteredElementCollector(doc)
            .OfCategory(BuiltInCategory.OST_Railings)
            .ToElements())

heights_mm = []
for rail in railings:
    p = rail.LookupParameter('Top Rail Height')
    heights_mm.append(p.AsDouble() * FT_TO_MM if p else None)

OUT = [list(railings), heights_mm]""",
        ))

        # Model text
        samples.append(_s(
            "Get all ModelText elements in the model using Dynamo Python",
            f"""\
{_REVIT_HEADER}

model_texts = (FilteredElementCollector(doc)
               .OfClass(ModelText)
               .ToElements())

texts = [mt.Text for mt in model_texts]

OUT = [list(model_texts), texts]""",
        ))

        # Filled regions
        samples.append(_s(
            "Get all FilledRegion elements in the active view using Dynamo Python",
            f"""\
{_REVIT_HEADER}

active_view = doc.ActiveView

filled = (FilteredElementCollector(doc, active_view.Id)
          .OfClass(FilledRegion)
          .ToElements())

OUT = list(filled)""",
        ))

        # Structural framing
        samples.append(_s(
            "Get all Structural Framing (beam/brace) family instances and their mark values using Dynamo Python",
            f"""\
{_REVIT_HEADER}

framing = (FilteredElementCollector(doc)
           .OfCategory(BuiltInCategory.OST_StructuralFraming)
           .OfClass(FamilyInstance)
           .ToElements())

marks = []
for f in framing:
    p = f.get_Parameter(BuiltInParameter.ALL_MODEL_MARK)
    marks.append(p.AsString() if p else '')

OUT = [list(framing), marks]""",
        ))

        # Plumbing fixtures
        samples.append(_s(
            "Collect all Plumbing Fixture instances in the model using Dynamo Python",
            f"""\
{_REVIT_HEADER}

fixtures = (FilteredElementCollector(doc)
            .OfCategory(BuiltInCategory.OST_PlumbingFixtures)
            .OfClass(FamilyInstance)
            .ToElements())

OUT = list(fixtures)""",
        ))

        # Electrical equipment
        samples.append(_s(
            "Collect all Electrical Equipment instances and return their circuit numbers using Dynamo Python",
            f"""\
{_REVIT_HEADER}

equipment = (FilteredElementCollector(doc)
             .OfCategory(BuiltInCategory.OST_ElectricalEquipment)
             .OfClass(FamilyInstance)
             .ToElements())

circuit_numbers = []
for eq in equipment:
    p = eq.LookupParameter('Panel')
    circuit_numbers.append(p.AsString() if p else None)

OUT = [list(equipment), circuit_numbers]""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- parameter manipulation
    # ------------------------------------------------------------------

    def _parameter_manipulation_extra(self) -> List[SAMPLE]:
        samples = []

        # Copy parameter value between elements
        samples.append(_s(
            "Copy the value of parameter IN[2] from element IN[0] to element IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

source = UnwrapElement(IN[0])
target = UnwrapElement(IN[1])
param_name = IN[2]

src_param = source.LookupParameter(param_name)
tgt_param = target.LookupParameter(param_name)

if src_param is not None and tgt_param is not None and not tgt_param.IsReadOnly:
    {_TX_START}
    if src_param.StorageType == StorageType.Double:
        tgt_param.Set(src_param.AsDouble())
    elif src_param.StorageType == StorageType.Integer:
        tgt_param.Set(src_param.AsInteger())
    elif src_param.StorageType == StorageType.String:
        tgt_param.Set(src_param.AsString())
    elif src_param.StorageType == StorageType.ElementId:
        tgt_param.Set(src_param.AsElementId())
    {_TX_END}

OUT = target""",
        ))

        # Read phase created
        samples.append(_s(
            "Get the Phase Created of a Revit element IN[0] as a string using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element = UnwrapElement(IN[0])

phase_id    = element.CreatedPhaseId
phase_elem  = doc.GetElement(phase_id)
phase_name  = phase_elem.Name if phase_elem is not None else None

OUT = phase_name""",
        ))

        # Read workset name
        samples.append(_s(
            "Get the Workset name of a Revit element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element = UnwrapElement(IN[0])

ws_param  = element.get_Parameter(BuiltInParameter.ELEM_PARTITION_PARAM)
ws_name   = ws_param.AsValueString() if ws_param is not None else None

OUT = ws_name""",
        ))

        # Set a YesNo parameter
        samples.append(_s(
            "Set a Yes/No parameter named IN[1] on element IN[0] to boolean value IN[2] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]
bool_value = bool(IN[2])

{_TX_START}

param = element.LookupParameter(param_name)
if param is not None and param.StorageType == StorageType.Integer:
    param.Set(1 if bool_value else 0)

{_TX_END}

OUT = element""",
        ))

        # Get type mark
        samples.append(_s(
            "Read the Type Mark parameter from the type of a FamilyInstance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])
symbol   = instance.Symbol

type_mark_param = symbol.get_Parameter(BuiltInParameter.ALL_MODEL_TYPE_MARK)
type_mark = type_mark_param.AsString() if type_mark_param is not None else None

OUT = type_mark""",
        ))

        # Compare parameter values across two elements
        samples.append(_s(
            "Compare the value of parameter IN[2] between elements IN[0] and IN[1] and return True if equal using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elem_a     = UnwrapElement(IN[0])
elem_b     = UnwrapElement(IN[1])
param_name = IN[2]

def get_val(elem, name):
    p = elem.LookupParameter(name)
    if p is None: return None
    if p.StorageType == StorageType.Double:  return p.AsDouble()
    if p.StorageType == StorageType.Integer: return p.AsInteger()
    return p.AsString()

val_a = get_val(elem_a, param_name)
val_b = get_val(elem_b, param_name)

are_equal = val_a == val_b

OUT = [are_equal, val_a, val_b]""",
        ))

        # Set elevation of a level
        samples.append(_s(
            "Set the elevation of a Level element IN[0] to IN[1] mm using Dynamo Python",
            f"""\
{_REVIT_HEADER}

MM_TO_FT = 1.0 / 304.8

level    = UnwrapElement(IN[0])
elev_mm  = float(IN[1])

{_TX_START}

level.Elevation = elev_mm * MM_TO_FT

{_TX_END}

OUT = level""",
        ))

        # Count elements with a parameter value
        samples.append(_s(
            "Count how many elements in list IN[0] have parameter IN[1] equal to value IN[2] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]
target_val = IN[2]

count = 0
for elem in elements:
    p = elem.LookupParameter(param_name)
    if p is None:
        continue
    if p.StorageType == StorageType.String:
        match = p.AsString() == str(target_val)
    elif p.StorageType == StorageType.Double:
        match = abs(p.AsDouble() - float(target_val)) < 1e-9
    elif p.StorageType == StorageType.Integer:
        match = p.AsInteger() == int(target_val)
    else:
        match = p.AsValueString() == str(target_val)
    if match:
        count += 1

OUT = count""",
        ))

        # Sum a parameter across a list
        samples.append(_s(
            "Sum the double value of parameter IN[1] across all elements in list IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]

total = 0.0
for elem in elements:
    p = elem.LookupParameter(param_name)
    if p is not None and p.StorageType == StorageType.Double:
        total += p.AsDouble()

OUT = total""",
        ))

        # Reset parameter to default
        samples.append(_s(
            "Reset (clear) a string parameter named IN[1] on element IN[0] to an empty string using Dynamo Python",
            f"""\
{_REVIT_HEADER}

element    = UnwrapElement(IN[0])
param_name = IN[1]

{_TX_START}

param = element.LookupParameter(param_name)
if param is not None and param.StorageType == StorageType.String and not param.IsReadOnly:
    param.Set('')

{_TX_END}

OUT = element""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- family instance operations
    # ------------------------------------------------------------------

    def _family_instance_operations_extra(self) -> List[SAMPLE]:
        samples = []

        # Activate all symbols in a family
        samples.append(_s(
            "Activate all FamilySymbols of a Family IN[0] so they can be placed using Dynamo Python",
            f"""\
{_REVIT_HEADER}

family = UnwrapElement(IN[0])

{_TX_START}

activated = []
for sym_id in family.GetFamilySymbolIds():
    sym = doc.GetElement(sym_id)
    if not sym.IsActive:
        sym.Activate()
        activated.append(sym)

{_TX_END}

OUT = activated""",
        ))

        # Get bounding box of instance
        samples.append(_s(
            "Get the bounding box of a FamilyInstance IN[0] in the active view and return min/max mm coordinates using Dynamo Python",
            f"""\
{_REVIT_HEADER}

FT_TO_MM = 304.8

instance    = UnwrapElement(IN[0])
active_view = doc.ActiveView

bbox = instance.get_BoundingBox(active_view)

if bbox is not None:
    min_mm = [bbox.Min.X * FT_TO_MM, bbox.Min.Y * FT_TO_MM, bbox.Min.Z * FT_TO_MM]
    max_mm = [bbox.Max.X * FT_TO_MM, bbox.Max.Y * FT_TO_MM, bbox.Max.Z * FT_TO_MM]
else:
    min_mm = max_mm = None

OUT = [min_mm, max_mm]""",
        ))

        # Flip facing orientation
        samples.append(_s(
            "Flip the facing orientation of a FamilyInstance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

{_TX_START}

if instance.CanFlipFacing:
    instance.flipFacing()

{_TX_END}

OUT = instance""",
        ))

        # Flip hand orientation
        samples.append(_s(
            "Flip the hand orientation of a FamilyInstance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

{_TX_START}

if instance.CanFlipHand:
    instance.flipHand()

{_TX_END}

OUT = instance""",
        ))

        # Group selection
        samples.append(_s(
            "Get all FamilyInstances inside a Group element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

group = UnwrapElement(IN[0])

member_ids = group.GetMemberIds()
instances  = [doc.GetElement(i) for i in member_ids
              if isinstance(doc.GetElement(i), FamilyInstance)]

OUT = instances""",
        ))

        # Get geometry solid
        samples.append(_s(
            "Extract the first Solid geometry from a FamilyInstance IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instance = UnwrapElement(IN[0])

options = Options()
options.DetailLevel = ViewDetailLevel.Fine

geom_elem = instance.get_Geometry(options)
solid = None
for g in geom_elem:
    if isinstance(g, Solid) and g.Volume > 0:
        solid = g
        break

OUT = solid""",
        ))

        # Rename instance via Mark
        samples.append(_s(
            "Set the Mark of each FamilyInstance in list IN[0] to a numbered prefix IN[1] followed by index using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instances = [UnwrapElement(e) for e in IN[0]]
prefix    = IN[1]

{_TX_START}

for i, inst in enumerate(instances):
    mark_param = inst.get_Parameter(BuiltInParameter.ALL_MODEL_MARK)
    if mark_param is not None and not mark_param.IsReadOnly:
        mark_param.Set(f"{{prefix}}{{i + 1:03d}}")

{_TX_END}

OUT = instances""",
        ))

        # Check if placed on correct level
        samples.append(_s(
            "Filter a list of FamilyInstances IN[0] to only those on Level IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

instances = [UnwrapElement(e) for e in IN[0]]
target_level = UnwrapElement(IN[1])

on_level = [inst for inst in instances if inst.LevelId == target_level.Id]

OUT = on_level""",
        ))

        # Place instance by face
        samples.append(_s(
            "Place a face-hosted FamilyInstance on the top face of a floor element IN[0] at its centroid using Dynamo Python",
            f"""\
{_REVIT_HEADER}

floor  = UnwrapElement(IN[0])
symbol = UnwrapElement(IN[1])

options      = Options()
options.DetailLevel = ViewDetailLevel.Fine
geom_elem    = floor.get_Geometry(options)

top_face = None
for solid in geom_elem:
    if isinstance(solid, Solid):
        for face in solid.Faces:
            if abs(face.FaceNormal.Z - 1.0) < 0.01:  # upward face
                top_face = face
                break

if top_face is not None:
    bbox      = top_face.GetBoundingBox()
    u_mid     = (bbox.Min.U + bbox.Max.U) / 2
    v_mid     = (bbox.Min.V + bbox.Max.V) / 2
    pt_on_face = top_face.Evaluate(UV(u_mid, v_mid))

    {_TX_START}

    if not symbol.IsActive:
        symbol.Activate()
        doc.Regenerate()

    instance = doc.Create.NewFamilyInstance(
        top_face.Reference,
        pt_on_face,
        XYZ.BasisX,
        symbol,
    )

    {_TX_END}

    OUT = instance
else:
    OUT = None""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- list operations
    # ------------------------------------------------------------------

    def _list_operations_extra(self) -> List[SAMPLE]:
        samples = []

        # Cartesian product
        samples.append(_s(
            "Compute the Cartesian product of two lists IN[0] and IN[1] in Dynamo Python",
            """\
list_a = IN[0]
list_b = IN[1]

product = [[a, b] for a in list_a for b in list_b]

OUT = product""",
        ))

        # Running total
        samples.append(_s(
            "Compute the running (cumulative) sum of a numeric list IN[0] in Dynamo Python",
            """\
numbers = IN[0]

running = []
total   = 0
for x in numbers:
    total += x
    running.append(total)

OUT = running""",
        ))

        # Max / min with index
        samples.append(_s(
            "Find the maximum and minimum values and their indices in a numeric list IN[0] using Dynamo Python",
            """\
numbers = IN[0]

max_val = max(numbers)
min_val = min(numbers)
max_idx = numbers.index(max_val)
min_idx = numbers.index(min_val)

OUT = [max_val, max_idx, min_val, min_idx]""",
        ))

        # Sliding window
        samples.append(_s(
            "Create a sliding window of size IN[1] over list IN[0] in Dynamo Python",
            """\
items = IN[0]
size  = int(IN[1])

windows = [items[i:i + size] for i in range(len(items) - size + 1)]

OUT = windows""",
        ))

        # Zip with index
        samples.append(_s(
            "Add an index to each item in a list IN[0] (like enumerate) in Dynamo Python",
            """\
items = IN[0]

indexed = [[i, item] for i, item in enumerate(items)]

OUT = indexed""",
        ))

        # Pairwise differences
        samples.append(_s(
            "Compute the pairwise differences between consecutive items in a numeric list IN[0] in Dynamo Python",
            """\
numbers = IN[0]

diffs = [numbers[i + 1] - numbers[i] for i in range(len(numbers) - 1)]

OUT = diffs""",
        ))

        # Normalize list
        samples.append(_s(
            "Normalize a list of numbers IN[0] to the range [0, 1] in Dynamo Python",
            """\
numbers = IN[0]

lo = min(numbers)
hi = max(numbers)
span = hi - lo

normalized = [(x - lo) / span if span != 0 else 0.0 for x in numbers]

OUT = normalized""",
        ))

        # Deep copy
        samples.append(_s(
            "Create a deep copy of a nested list IN[0] in Dynamo Python",
            """\
import copy

nested = IN[0]
copied = copy.deepcopy(nested)

OUT = copied""",
        ))

        # Zip lists of unequal length (fill with None)
        samples.append(_s(
            "Zip two lists IN[0] and IN[1] of potentially unequal length, filling missing values with None in Dynamo Python",
            """\
from itertools import zip_longest

list_a = IN[0]
list_b = IN[1]

padded = [[a, b] for a, b in zip_longest(list_a, list_b, fillvalue=None)]

OUT = padded""",
        ))

        # List statistics
        samples.append(_s(
            "Compute mean, median, and standard deviation of a numeric list IN[0] in Dynamo Python",
            """\
import statistics

numbers = IN[0]

mean   = statistics.mean(numbers)
median = statistics.median(numbers)
stdev  = statistics.stdev(numbers) if len(numbers) > 1 else 0.0

OUT = [mean, median, stdev]""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- coordinate transforms
    # ------------------------------------------------------------------

    def _coordinate_transforms_extra(self) -> List[SAMPLE]:
        samples = []

        # Scale non-uniformly
        samples.append(_s(
            "Scale a geometry object IN[0] non-uniformly by (IN[1], IN[2], IN[3]) factors along X, Y, Z in Dynamo",
            f"""\
{_PROTO_HEADER}

geom = IN[0]
sx   = float(IN[1])
sy   = float(IN[2])
sz   = float(IN[3])

origin   = Point.ByCoordinates(0, 0, 0)
cs_from  = CoordinateSystem.Identity()

# Non-uniform scale: build target CS with scaled axis vectors
cs_to = CoordinateSystem.ByOriginVectors(
    origin,
    Vector.ByCoordinates(sx, 0, 0),
    Vector.ByCoordinates(0,  sy, 0),
)
# Z scaling applied via transform
scaled = geom.Transform(cs_from, cs_to)

OUT = scaled""",
        ))

        # Align two coordinate systems
        samples.append(_s(
            "Compute the Transform needed to align coordinate system IN[0] to coordinate system IN[1] in Dynamo Python",
            f"""\
{_REVIT_HEADER}

cs_from = IN[0]  # Revit Transform or CoordinateSystem
cs_to   = IN[1]

# For Revit Transform objects:
# The relative transform = cs_to * cs_from.Inverse
relative = cs_to.Multiply(cs_from.Inverse)

OUT = relative""",
        ))

        # Convert Dynamo point to Revit XYZ
        samples.append(_s(
            "Convert a Dynamo Point IN[0] to a Revit XYZ object using Dynamo Python",
            f"""\
{_REVIT_HEADER}
{_REVIT_NODES_HEADER}

dynamo_pt = IN[0]

revit_xyz = dynamo_pt.ToRevitType()

OUT = revit_xyz""",
        ))

        # Convert Revit XYZ to Dynamo Point
        samples.append(_s(
            "Convert a Revit XYZ IN[0] to a Dynamo Point using Dynamo Python",
            f"""\
{_REVIT_HEADER}
{_REVIT_NODES_HEADER}

revit_xyz = IN[0]

dynamo_pt = revit_xyz.ToPoint()

OUT = dynamo_pt""",
        ))

        # Convert Revit curve to Dynamo curve
        samples.append(_s(
            "Convert a Revit DB Curve IN[0] to a Dynamo Curve using Dynamo Python",
            f"""\
{_REVIT_HEADER}
{_REVIT_NODES_HEADER}

revit_curve = IN[0]

dynamo_curve = revit_curve.ToProtoType()

OUT = dynamo_curve""",
        ))

        # Barycentric interpolation
        samples.append(_s(
            "Linearly interpolate between two points IN[0] and IN[1] at parameter IN[2] (0-1) in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

pt_a = IN[0]
pt_b = IN[1]
t    = float(IN[2])

# Vector from A to B, scaled by t
v_ab     = Vector.ByTwoPoints(pt_a, pt_b)
scaled_v = v_ab.Scale(t)
result   = pt_a.Translate(scaled_v)

OUT = result""",
        ))

        # World to local coordinates
        samples.append(_s(
            "Transform a world-space Point IN[0] into local coordinates of CoordinateSystem IN[1] in Dynamo",
            f"""\
{_PROTO_HEADER}

world_pt = IN[0]
local_cs = IN[1]  # CoordinateSystem

# Transform from world (Identity) to local CS
identity = CoordinateSystem.Identity()
local_pt = world_pt.Transform(identity, local_cs)

OUT = local_pt""",
        ))

        # Arc from center, radius, angles
        samples.append(_s(
            "Create an Arc centered at origin with radius IN[0] mm from start angle IN[1] to end angle IN[2] degrees in Dynamo",
            f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

radius_mm   = IN[0]
start_deg   = IN[1]
end_deg     = IN[2]

radius      = radius_mm * MM_TO_FT
start_rad   = math.radians(start_deg)
end_rad     = math.radians(end_deg)
sweep_angle = end_rad - start_rad

cs = CoordinateSystem.Identity()

arc = Arc.ByCenterPointRadiusAngle(
    Point.ByCoordinates(0, 0, 0),
    radius,
    start_deg,
    end_deg,
    Vector.ZAxis(),
)

OUT = arc""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- curve operations
    # ------------------------------------------------------------------

    def _curve_operations_extra(self) -> List[SAMPLE]:
        samples = []

        # Curve area (approximation)
        samples.append(_s(
            "Approximate the area enclosed by a closed planar curve IN[0] using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

curve = IN[0]  # closed curve

surface = Surface.ByPatch(curve)
area_ft2 = surface.Area
area_mm2 = area_ft2 * (304.8 ** 2)

surface.Dispose()

OUT = area_mm2""",
        ))

        # Map points to curve
        samples.append(_s(
            "Project a list of points IN[0] onto a curve IN[1] and return the closest points on the curve in Dynamo",
            f"""\
{_PROTO_HEADER}

points = IN[0]
curve  = IN[1]

projected = [curve.ClosestPointTo(pt) for pt in points]
params    = [curve.ParameterAtPoint(p) for p in projected]

OUT = [projected, params]""",
        ))

        # Resample curve
        samples.append(_s(
            "Resample a curve IN[0] as IN[1] equally-spaced points along its length in Dynamo",
            f"""\
{_PROTO_HEADER}

curve = IN[0]
n     = int(IN[1])

pts = [curve.PointAtParameter(i / (n - 1)) for i in range(n)]

OUT = pts""",
        ))

        # Pipe along path
        samples.append(_s(
            "Create a pipe solid of radius IN[0] mm along a path curve IN[1] using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

radius_mm = IN[0]
path      = IN[1]

start_pt = path.StartPoint
tangent  = path.TangentAtParameter(0)

profile = Circle.ByCenterPointRadiusNormal(start_pt, radius_mm * MM_TO_FT, tangent)

solid = Solid.BySweep(profile, path)

OUT = solid""",
        ))

        # Spiral in 3D
        samples.append(_s(
            "Create a 3D Archimedean spiral starting at radius IN[0] mm growing to IN[1] mm with IN[2] turns in Dynamo Python",
            f"""\
{_PROTO_HEADER}
import math

MM_TO_FT = 1.0 / 304.8

r_start_mm = IN[0]
r_end_mm   = IN[1]
turns      = IN[2]

pts_per_turn = 72
total_pts    = int(turns * pts_per_turn) + 1

pts = []
for i in range(total_pts):
    t     = i / (total_pts - 1)
    angle = 2 * math.pi * turns * t
    r     = (r_start_mm + (r_end_mm - r_start_mm) * t) * MM_TO_FT
    pts.append(Point.ByCoordinates(
        r * math.cos(angle),
        r * math.sin(angle),
        0,
    ))

spiral = NurbsCurve.ByPoints(pts)

OUT = spiral""",
        ))

        # Blend two curves
        samples.append(_s(
            "Create a smooth blending curve between the end of curve IN[0] and the start of curve IN[1] in Dynamo",
            f"""\
{_PROTO_HEADER}

curve_a = IN[0]
curve_b = IN[1]

end_pt   = curve_a.EndPoint
start_pt = curve_b.StartPoint

# Control points for G1 blend
tan_a = curve_a.TangentAtParameter(1.0)
tan_b = curve_b.TangentAtParameter(0.0)

mid_a = end_pt.Translate(tan_a.Scale(0.3))
mid_b = start_pt.Translate(tan_b.Reverse().Scale(0.3))

blend = NurbsCurve.ByControlPoints([end_pt, mid_a, mid_b, start_pt], degree=3)

OUT = blend""",
        ))

        # Evaluate frenet frame
        samples.append(_s(
            "Compute the Frenet frame (tangent, normal, binormal) at parameter IN[1] on curve IN[0] in Dynamo Python",
            f"""\
{_PROTO_HEADER}

curve = IN[0]
t     = float(IN[1])

tangent   = curve.TangentAtParameter(t)
normal    = curve.NormalAtParameter(t)
binormal  = tangent.Cross(normal)

OUT = [tangent, normal, binormal]""",
        ))

        # Length segments
        samples.append(_s(
            "Split a curve IN[0] into segments of equal length IN[1] mm in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

curve      = IN[0]
seg_len_mm = float(IN[1])
seg_len_ft = seg_len_mm * MM_TO_FT

total_len = curve.Length
n_segs    = max(1, int(total_len / seg_len_ft))

params    = [i / n_segs for i in range(n_segs + 1)]
pts       = [curve.PointAtParameter(t) for t in params]
segments  = [Line.ByStartPointEndPoint(pts[i], pts[i + 1]) for i in range(n_segs)]

OUT = segments""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- surface operations
    # ------------------------------------------------------------------

    def _surface_operations_extra(self) -> List[SAMPLE]:
        samples = []

        # Surface by four points
        samples.append(_s(
            "Create a Surface through four corner points IN[0], IN[1], IN[2], IN[3] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

p0 = IN[0]
p1 = IN[1]
p2 = IN[2]
p3 = IN[3]

boundary = PolyCurve.ByPoints([p0, p1, p2, p3], connectLastToFirst=True)
surface  = Surface.ByPatch(boundary)

OUT = surface""",
        ))

        # Trim surface
        samples.append(_s(
            "Trim a surface IN[0] with a cutting surface IN[1] and return the remaining pieces in Dynamo",
            f"""\
{_PROTO_HEADER}

surface = IN[0]
cutter  = IN[1]

trimmed_pieces = surface.Trim(cutter)

OUT = list(trimmed_pieces)""",
        ))

        # Surface isolines
        samples.append(_s(
            "Extract IN[1] isocurves at equal V-parameter spacing from surface IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

surface = IN[0]
count   = int(IN[1])

iso_curves = []
for i in range(count):
    v = i / (count - 1) if count > 1 else 0.5
    curve = surface.GetIsoline(1, v)  # direction 1 = V iso
    iso_curves.append(curve)

OUT = iso_curves""",
        ))

        # Surface normal vectors as arrows
        samples.append(_s(
            "Sample a grid of IN[1] x IN[2] normal vectors on surface IN[0] as lines in Dynamo Python",
            f"""\
{_PROTO_HEADER}

surface = IN[0]
u_count = int(IN[1])
v_count = int(IN[2])
scale   = IN[3] if IN[3] else 0.1  # arrow length in feet

arrows = []
for i in range(u_count):
    u = i / (u_count - 1) if u_count > 1 else 0.5
    for j in range(v_count):
        v = j / (v_count - 1) if v_count > 1 else 0.5
        pt = surface.PointAtParameter(u, v)
        n  = surface.NormalAtParameter(u, v)
        tip = pt.Translate(n.Scale(scale))
        arrows.append(Line.ByStartPointEndPoint(pt, tip))

OUT = arrows""",
        ))

        # Developable surface check via Gaussian curvature
        samples.append(_s(
            "Compute Gaussian curvature at a UV point IN[1], IN[2] on surface IN[0] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

surface = IN[0]
u       = float(IN[1])
v       = float(IN[2])

# PrincipalCurvatures returns [k1, k2] at the UV point
curvatures = surface.PrincipalCurvatures(u, v)
gaussian   = curvatures[0] * curvatures[1]

OUT = [curvatures[0], curvatures[1], gaussian]""",
        ))

        # Multi-loft with guides
        samples.append(_s(
            "Create a guided loft surface through profiles IN[0] with guide curves IN[1] in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

profiles = IN[0]  # list of curves
guides   = IN[1]  # list of guide curves

surface = Surface.ByLoft(profiles, guides)

OUT = surface""",
        ))

        # Floor surface from room boundary
        samples.append(_s(
            "Create a Surface from the boundary curves of a Room element IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}

room = UnwrapElement(IN[0])

# Get room boundary segments
options   = SpatialElementBoundaryOptions()
seg_loops = room.GetBoundarySegments(options)

import clr
clr.AddReference('ProtoGeometry')
from Autodesk.DesignScript.Geometry import *
{_REVIT_NODES_HEADER}

surfaces = []
for loop in seg_loops:
    pts = []
    for seg in loop:
        curve = seg.GetCurve()
        pts.append(curve.GetEndPoint(0).ToPoint())
    if pts:
        poly    = PolyCurve.ByPoints(pts, connectLastToFirst=True)
        surface = Surface.ByPatch(poly)
        surfaces.append(surface)

OUT = surfaces""",
        ))

        # Surface volume
        samples.append(_s(
            "Compute the volume of a solid IN[0] in mm^3 using Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

solid = IN[0]

FT3_TO_MM3 = 304.8 ** 3

volume_ft3 = solid.Volume
volume_mm3 = volume_ft3 * FT3_TO_MM3

OUT = volume_mm3""",
        ))

        # Offset surface inward
        samples.append(_s(
            "Shell a solid IN[0] with wall thickness IN[1] mm (keep outer, hollow interior) in Dynamo ProtoGeometry",
            f"""\
{_PROTO_HEADER}

MM_TO_FT = 1.0 / 304.8

solid    = IN[0]
thick_mm = IN[1]

faces_to_remove = []  # empty = all faces kept; add face indices to open faces
shell = solid.Shell(faces_to_remove, thick_mm * MM_TO_FT)

OUT = shell""",
        ))

        return samples

    # ------------------------------------------------------------------
    # Extra samples -- export data
    # ------------------------------------------------------------------

    def _export_data_extra(self) -> List[SAMPLE]:
        samples = []

        # Export to TSV
        samples.append(_s(
            "Write a 2D list IN[0] to a tab-separated (.tsv) file at path IN[1] using Dynamo Python",
            """\
file_path = IN[1]
data      = IN[0]

lines = ['\\t'.join(str(v) for v in row) for row in data]

with open(file_path, 'w', encoding='utf-8') as f:
    f.write('\\n'.join(lines))

OUT = file_path""",
        ))

        # Read schedule and write to Excel
        samples.append(_s(
            "Read a Revit ViewSchedule IN[0] and write its data to Excel at path IN[1] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import openpyxl

schedule  = UnwrapElement(IN[0])
file_path = IN[1]

table  = schedule.GetTableData()
body   = table.GetSectionData(SectionType.Body)
header = table.GetSectionData(SectionType.Header)

n_cols  = body.NumberOfColumns
n_rows  = body.NumberOfRows
h_rows  = header.NumberOfRows

wb = openpyxl.Workbook()
ws = wb.active
ws.title = schedule.Name[:31]  # Excel sheet names <= 31 chars

for r in range(h_rows):
    ws.append([schedule.GetCellText(SectionType.Header, r, c) for c in range(n_cols)])

for r in range(n_rows):
    ws.append([schedule.GetCellText(SectionType.Body, r, c) for c in range(n_cols)])

wb.save(file_path)

OUT = file_path""",
        ))

        # Export material quantities
        samples.append(_s(
            "Export each material name and total volume/area from all elements to CSV at IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

FT3_TO_M3 = 0.0283168
file_path = IN[0]

elements = (FilteredElementCollector(doc)
            .WhereElementIsNotElementType()
            .ToElements())

mat_volumes = {{}}
for elem in elements:
    try:
        mats = elem.GetMaterialIds(False)
    except Exception:
        continue
    for mat_id in mats:
        mat   = doc.GetElement(mat_id)
        if mat is None:
            continue
        vol_ft3 = elem.GetMaterialVolume(mat_id)
        mat_volumes[mat.Name] = mat_volumes.get(mat.Name, 0) + vol_ft3

header = ['MaterialName', 'Volume_m3']
rows   = [header] + [[name, vol * FT3_TO_M3] for name, vol in sorted(mat_volumes.items())]

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(rows)

OUT = file_path""",
        ))

        # Batch write JSON records
        samples.append(_s(
            "Write each Revit element in IN[0] as a JSON object with its ElementId and parameter IN[1] to a JSONL file IN[2] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import json

elements   = [UnwrapElement(e) for e in IN[0]]
param_name = IN[1]
file_path  = IN[2]

with open(file_path, 'w', encoding='utf-8') as f:
    for elem in elements:
        p   = elem.LookupParameter(param_name)
        val = p.AsValueString() if p else None
        record = {{'ElementId': elem.Id.IntegerValue, param_name: val}}
        f.write(json.dumps(record) + '\\n')

OUT = file_path""",
        ))

        # Export door schedule
        samples.append(_s(
            "Export all Door instances with their Mark, Width, Height, and Host Wall Type to CSV at IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

FT_TO_MM = 304.8
file_path = IN[0]

doors = (FilteredElementCollector(doc)
         .OfCategory(BuiltInCategory.OST_Doors)
         .OfClass(FamilyInstance)
         .ToElements())

header = ['ElementId', 'Mark', 'FamilyType', 'Width_mm', 'Height_mm', 'HostWallType']
rows   = [header]
for door in doors:
    eid       = door.Id.IntegerValue
    mark_p    = door.get_Parameter(BuiltInParameter.ALL_MODEL_MARK)
    mark      = mark_p.AsString() if mark_p else ''
    sym_name  = door.Symbol.Name
    w_p       = door.Symbol.get_Parameter(BuiltInParameter.DOOR_WIDTH)
    h_p       = door.Symbol.get_Parameter(BuiltInParameter.DOOR_HEIGHT)
    w_mm      = w_p.AsDouble() * FT_TO_MM if w_p else ''
    h_mm      = h_p.AsDouble() * FT_TO_MM if h_p else ''
    host_type = doc.GetElement(door.Host.GetTypeId()).Name if door.Host else ''
    rows.append([eid, mark, sym_name, w_mm, h_mm, host_type])

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(rows)

OUT = file_path""",
        ))

        # Read CSV as dict
        samples.append(_s(
            "Read a CSV with a header row at path IN[0] and return a list of dictionaries (one per row) using Dynamo Python",
            """\
import csv

file_path = IN[0]

with open(file_path, 'r', newline='', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    records = [dict(row) for row in reader]

keys = list(records[0].keys()) if records else []

OUT = [records, keys]""",
        ))

        # Export view list
        samples.append(_s(
            "Export all views with their names, types, and sheet placements to CSV at IN[0] using Dynamo Python",
            f"""\
{_REVIT_HEADER}
import csv

file_path = IN[0]

views = (FilteredElementCollector(doc)
         .OfClass(View)
         .ToElements())

# Map views to sheets
sheet_map = {{}}
viewports = (FilteredElementCollector(doc)
             .OfClass(Viewport)
             .ToElements())
for vp in viewports:
    view_id   = vp.ViewId
    sheet     = doc.GetElement(vp.SheetId)
    sheet_map[view_id.IntegerValue] = sheet.SheetNumber if sheet else ''

header = ['ElementId', 'ViewName', 'ViewType', 'SheetNumber']
rows   = [header]
for v in views:
    if v.IsTemplate:
        continue
    rows.append([
        v.Id.IntegerValue,
        v.Name,
        str(v.ViewType),
        sheet_map.get(v.Id.IntegerValue, ''),
    ])

with open(file_path, 'w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(rows)

OUT = file_path""",
        ))

        return samples
